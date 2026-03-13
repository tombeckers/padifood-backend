"""
Invoice verification script for Padifood / OTTO Workforce.

Compares:
1. Hours on OTTO's invoice (Padifood specificatie) vs Padifood's own timesheets (Kloklijst)
2. Rates on OTTO's invoice vs OTTO's agreed rate card per employee

Outputs discrepancies to output/verification_report.xlsx
"""

import csv
import os
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment


# ---------------------------------------------------------------------------
# Config
# ---------------------------------------------------------------------------
INVOICE_FILE = "formatted_input/Padifood specificatie - Export Factuur.csv"
KLOKLIJST_FILE = "formatted_input/202551 Kloklijst Padifood Otto Workforce.csv"
OTTO_RATES_FILE = "formatted_input/OTTO -Padifood tarievenoverzicht per persoon - achternaam voornaam.csv"
OUTPUT_DIR = "output"
OUTPUT_FILE = os.path.join(OUTPUT_DIR, "verification_report.xlsx")

# Hour type columns shared between kloklijst and invoice
HOUR_TYPES = [
    "Norm uren Dag",
    "T133 Dag",
    "T135 Dag",
    "T200 Dag",
    "OW140 Week",
    "OW180 Dag",
    "OW200 Dag",
]

# Map invoice hour code to the rate multiplier column index in OTTO rates file
# OTTO rates header: Personeelsnummer,Achternaam,Voornaam,Wekenteller,Schaal,
#   Uurloon incl. ATV, Uurloon excl. ATV, 1, 1.33, 1.35, 1.8, 2, 1.4, 1.8, 2
# Indices (0-based after DictReader): cols 7-15 map to multipliers
RATE_COLUMN_NAMES = ["1", "1.33", "1.35", "1.8", "2", "1.4", "1.8", "2"]

# Map invoice "Code toeslag" to the expected rate column in OTTO rates
CODE_TO_RATE_INDEX = {
    "Norm uren Dag": 0,   # 1x
    "T133 Dag": 1,        # 1.33x
    "T135 Dag": 2,        # 1.35x
    "OW180 Dag": 3,       # 1.8x day
    "T200 Dag": 4,        # 2x day
    "OW140 Week": 5,      # 1.4x overtime week
    "OW180 Dag_ow": 3,    # 1.8x (same column for day)
    "OW200 Dag": 7,       # 2x overtime
}

# We need to handle the duplicate column names (1.8 and 2 appear twice).
# In the CSV the columns after the header are positional. Let's read by position.


def safe_float(val):
    """Convert string to float, return 0.0 on failure."""
    try:
        return float(val.replace(",", ".")) if val else 0.0
    except (ValueError, AttributeError):
        return 0.0


def normalize_name(name):
    """Normalize a name for matching: lowercase, sorted words."""
    return " ".join(sorted(name.lower().replace("-", " ").split()))


# ---------------------------------------------------------------------------
# 1. Load OTTO rate card (positional read for duplicate column names)
# ---------------------------------------------------------------------------
def load_otto_rates():
    """Returns dict: personeelsnummer -> {name, schaal, rates_by_position}
    and dict: normalized_name -> personeelsnummer
    """
    rates = {}
    name_to_id = {}
    with open(OTTO_RATES_FILE, encoding="utf-8-sig") as f:
        reader = csv.reader(f)
        header = next(reader)
        # Positions: 0=Personeelsnummer, 1=Achternaam, 2=Voornaam, 3=Wekenteller,
        # 4=Schaal, 5=Uurloon incl ATV, 6=Uurloon excl ATV,
        # 7=1x, 8=1.33x, 9=1.35x, 10=1.8x(day), 11=2x(day),
        # 12=1.4x(ow), 13=1.8x(ow), 14=2x(ow)
        for row in reader:
            if len(row) < 15:
                continue
            pid = row[0].strip()
            achternaam = row[1].strip()
            voornaam = row[2].strip()
            if not achternaam and not voornaam:
                continue
            full_name = f"{voornaam} {achternaam}".strip()
            schaal = row[4].strip()
            rate_values = [safe_float(row[i]) for i in range(7, 15)]
            uurloon_incl = safe_float(row[5])
            uurloon_excl = safe_float(row[6])

            entry = {
                "name": full_name,
                "achternaam": achternaam,
                "voornaam": voornaam,
                "schaal": schaal,
                "uurloon_incl": uurloon_incl,
                "uurloon_excl": uurloon_excl,
                "rates": rate_values,  # [1x, 1.33x, 1.35x, 1.8x_day, 2x_day, 1.4x_ow, 1.8x_ow, 2x_ow]
            }
            if pid:
                rates[pid] = entry
            name_to_id[normalize_name(full_name)] = pid or full_name
        return rates, name_to_id


# ---------------------------------------------------------------------------
# 2. Load Kloklijst (timesheets)
# ---------------------------------------------------------------------------
def load_kloklijst():
    """Returns dict: normalized_name -> {date_str -> {hour_type -> hours}}"""
    timesheets = defaultdict(lambda: defaultdict(lambda: defaultdict(float)))
    current_name = None
    current_pid = None

    with open(KLOKLIJST_FILE, encoding="utf-8-sig") as f:
        reader = csv.reader(f)
        header = next(reader)
        # Find column indices
        col = {h.strip(): i for i, h in enumerate(header)}

        for row in reader:
            if len(row) < len(header) - 1:
                continue

            name = row[col["Naam"]].strip() if col.get("Naam") is not None else ""
            if name:
                current_name = name
                current_pid = row[col.get("Personeelsnummer", 1)].strip()

            date_str = row[col["Datum"]].strip() if col.get("Datum") is not None else ""
            if not date_str or not current_name:
                continue

            date_key = date_str[:10]  # YYYY-MM-DD

            # Skip summary rows (rows where Datum is empty but have totals)
            # These are the rows without a date - already skipped above

            for ht in HOUR_TYPES:
                if ht in col:
                    val = safe_float(row[col[ht]])
                    if val != 0:
                        timesheets[normalize_name(current_name)][date_key][ht] += val

    return timesheets


# ---------------------------------------------------------------------------
# 3. Load Invoice
# ---------------------------------------------------------------------------
def load_invoice():
    """Returns list of invoice line items and aggregated hours per employee/date/type."""
    lines = []
    agg = defaultdict(lambda: defaultdict(lambda: defaultdict(float)))

    with open(INVOICE_FILE, encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        for row in reader:
            sap_id = row["SAP ID"].strip()
            name = row["Naam"].strip()
            date_key = row["Datum"].strip()[:10]
            code = row["Code toeslag"].strip()
            hours = safe_float(row["Totaal uren"])
            subtotal = safe_float(row["Subtotaal"])
            uurloon = safe_float(row["Uurloon"])
            uurloon_excl = safe_float(row["Uurloon zonder ATV"])
            wekentelling = row["Wekentelling"].strip()
            fase = row["Fase tarief"].strip()

            rate_per_hour = subtotal / hours if hours else 0

            lines.append({
                "sap_id": sap_id,
                "name": name,
                "date": date_key,
                "code": code,
                "hours": hours,
                "subtotal": subtotal,
                "rate_per_hour": rate_per_hour,
                "uurloon": uurloon,
                "uurloon_excl": uurloon_excl,
                "wekentelling": wekentelling,
                "fase": fase,
            })

            norm_name = normalize_name(name)
            agg[norm_name][date_key][code] += hours

    return lines, agg


# ---------------------------------------------------------------------------
# 4. Compare hours
# ---------------------------------------------------------------------------
def compare_hours(invoice_agg, kloklijst):
    """Compare aggregated hours. Returns list of discrepancies."""
    discrepancies = []

    # Map invoice code to kloklijst column name
    code_to_klok = {
        "Norm uren Dag": "Norm uren Dag",
        "T133 Dag": "T133 Dag",
        "T135 Dag": "T135 Dag",
        "T200 Dag": "T200 Dag",
        "OW140 Week": "OW140 Week",
        "OW180 Dag": "OW180 Dag",
        "OW200 Dag": "OW200 Dag",
    }

    all_names = set(invoice_agg.keys()) | set(kloklijst.keys())

    for norm_name in sorted(all_names):
        inv_dates = invoice_agg.get(norm_name, {})
        klok_dates = kloklijst.get(norm_name, {})
        all_dates = set(inv_dates.keys()) | set(klok_dates.keys())

        for date in sorted(all_dates):
            inv_codes = inv_dates.get(date, {})
            klok_codes = klok_dates.get(date, {})
            all_codes = set(inv_codes.keys()) | set(code_to_klok.keys())

            for code in sorted(all_codes):
                klok_col = code_to_klok.get(code, code)
                inv_hours = inv_codes.get(code, 0)
                klok_hours = klok_codes.get(klok_col, 0)

                # Only report if at least one side has hours
                if inv_hours == 0 and klok_hours == 0:
                    continue

                diff = round(inv_hours - klok_hours, 4)
                if abs(diff) > 0.001:
                    discrepancies.append({
                        "name": norm_name,
                        "date": date,
                        "hour_type": code,
                        "invoice_hours": inv_hours,
                        "kloklijst_hours": klok_hours,
                        "difference": diff,
                    })

    return discrepancies


# ---------------------------------------------------------------------------
# 5. Compare rates
# ---------------------------------------------------------------------------
def compare_rates(invoice_lines, otto_rates, name_to_id):
    """Compare invoiced rate per hour vs OTTO agreed rate. Returns discrepancies."""
    rate_issues = []

    # Rate index by invoice code
    # rates array: [1x, 1.33x, 1.35x, 1.8x_day, 2x_day, 1.4x_ow, 1.8x_ow, 2x_ow]
    code_rate_idx = {
        "Norm uren Dag": 0,   # 1x
        "T133 Dag": 1,        # 1.33x
        "T135 Dag": 2,        # 1.35x
        "T200 Dag": 4,        # 2x day
        "OW140 Week": 5,      # 1.4x overtime
        "OW180 Dag": 6,       # 1.8x overtime (NOT index 3 which is day 1.8x)
        "OW200 Dag": 7,       # 2x overtime
    }

    checked = set()

    for line in invoice_lines:
        norm_name = normalize_name(line["name"])
        code = line["code"]
        key = (norm_name, code)
        if key in checked:
            continue
        checked.add(key)

        # Find in OTTO rates
        pid = name_to_id.get(norm_name)
        if not pid or pid not in otto_rates:
            rate_issues.append({
                "sap_id": line["sap_id"],
                "name": line["name"],
                "hour_type": code,
                "invoice_rate": line["rate_per_hour"],
                "otto_rate": "NOT FOUND",
                "difference": "N/A",
                "invoice_uurloon": line["uurloon"],
                "otto_schaal": "N/A",
            })
            continue

        otto = otto_rates[pid]
        idx = code_rate_idx.get(code)
        if idx is None:
            continue

        expected_rate = otto["rates"][idx]
        actual_rate = line["rate_per_hour"]
        diff = round(actual_rate - expected_rate, 2)

        if abs(diff) > 0.02:  # allow 2 cent rounding tolerance
            rate_issues.append({
                "sap_id": line["sap_id"],
                "name": line["name"],
                "hour_type": code,
                "invoice_rate": round(actual_rate, 4),
                "otto_rate": round(expected_rate, 4),
                "difference": diff,
                "invoice_uurloon": line["uurloon"],
                "otto_schaal": otto["schaal"],
            })

    return rate_issues


# ---------------------------------------------------------------------------
# 6. Build summary per employee
# ---------------------------------------------------------------------------
def build_employee_summary(invoice_lines, kloklijst, otto_rates, name_to_id):
    """Per-employee summary: total hours invoiced vs kloklijst, total amount."""
    employees = defaultdict(lambda: {
        "sap_id": "",
        "invoice_hours": 0,
        "kloklijst_hours": 0,
        "invoice_total": 0,
        "schaal": "",
        "in_kloklijst": False,
        "in_otto_rates": False,
    })

    for line in invoice_lines:
        norm_name = normalize_name(line["name"])
        e = employees[line["name"]]
        e["sap_id"] = line["sap_id"]
        e["invoice_hours"] += line["hours"]
        e["invoice_total"] += line["subtotal"]
        pid = name_to_id.get(norm_name)
        if pid and pid in otto_rates:
            e["schaal"] = otto_rates[pid]["schaal"]
            e["in_otto_rates"] = True

    for norm_name, dates in kloklijst.items():
        total = 0
        for date_codes in dates.values():
            for hours in date_codes.values():
                total += hours
        # Find the original name from invoice
        for orig_name in employees:
            if normalize_name(orig_name) == norm_name:
                employees[orig_name]["kloklijst_hours"] = total
                employees[orig_name]["in_kloklijst"] = True
                break
        else:
            # Employee in kloklijst but not in invoice - skip for now
            pass

    return employees


# ---------------------------------------------------------------------------
# 7. Write Excel output
# ---------------------------------------------------------------------------
def write_report(hour_disc, rate_disc, employee_summary):
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    wb = Workbook()

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    warn_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    def write_header(ws, headers):
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

    # --- Sheet 1: Employee Summary ---
    ws1 = wb.active
    ws1.title = "Employee Summary"
    headers = ["SAP ID", "Name", "Schaal", "Invoice Hours", "Kloklijst Hours",
               "Hours Difference", "Invoice Total (€)", "In Kloklijst", "In OTTO Rates"]
    write_header(ws1, headers)

    for row_idx, (name, e) in enumerate(sorted(employee_summary.items()), 2):
        diff = round(e["invoice_hours"] - e["kloklijst_hours"], 2)
        ws1.cell(row=row_idx, column=1, value=e["sap_id"])
        ws1.cell(row=row_idx, column=2, value=name)
        ws1.cell(row=row_idx, column=3, value=e["schaal"])
        ws1.cell(row=row_idx, column=4, value=round(e["invoice_hours"], 2))
        ws1.cell(row=row_idx, column=5, value=round(e["kloklijst_hours"], 2))
        diff_cell = ws1.cell(row=row_idx, column=6, value=diff)
        if abs(diff) > 0.01:
            diff_cell.fill = warn_fill
        ws1.cell(row=row_idx, column=7, value=round(e["invoice_total"], 2))
        klok_cell = ws1.cell(row=row_idx, column=8, value="Yes" if e["in_kloklijst"] else "NO")
        if not e["in_kloklijst"]:
            klok_cell.fill = warn_fill
        otto_cell = ws1.cell(row=row_idx, column=9, value="Yes" if e["in_otto_rates"] else "NO")
        if not e["in_otto_rates"]:
            otto_cell.fill = warn_fill

    for col in range(1, len(headers) + 1):
        ws1.column_dimensions[chr(64 + col) if col <= 26 else "A"].width = 18

    # --- Sheet 2: Hour Discrepancies ---
    ws2 = wb.create_sheet("Hour Discrepancies")
    headers2 = ["Name", "Date", "Hour Type", "Invoice Hours", "Kloklijst Hours", "Difference"]
    write_header(ws2, headers2)

    for row_idx, d in enumerate(hour_disc, 2):
        ws2.cell(row=row_idx, column=1, value=d["name"])
        ws2.cell(row=row_idx, column=2, value=d["date"])
        ws2.cell(row=row_idx, column=3, value=d["hour_type"])
        ws2.cell(row=row_idx, column=4, value=d["invoice_hours"])
        ws2.cell(row=row_idx, column=5, value=d["kloklijst_hours"])
        diff_cell = ws2.cell(row=row_idx, column=6, value=d["difference"])
        diff_cell.fill = warn_fill

    # --- Sheet 3: Rate Discrepancies ---
    ws3 = wb.create_sheet("Rate Discrepancies")
    headers3 = ["SAP ID", "Name", "Schaal", "Hour Type", "Invoice Rate (€/hr)",
                "OTTO Agreed Rate (€/hr)", "Difference (€)", "Invoice Uurloon"]
    write_header(ws3, headers3)

    for row_idx, d in enumerate(rate_disc, 2):
        ws3.cell(row=row_idx, column=1, value=d["sap_id"])
        ws3.cell(row=row_idx, column=2, value=d["name"])
        ws3.cell(row=row_idx, column=3, value=d["otto_schaal"])
        ws3.cell(row=row_idx, column=4, value=d["hour_type"])
        ws3.cell(row=row_idx, column=5, value=d["invoice_rate"])
        ws3.cell(row=row_idx, column=6, value=d["otto_rate"])
        ws3.cell(row=row_idx, column=7, value=d["difference"])
        ws3.cell(row=row_idx, column=8, value=d["invoice_uurloon"])

    wb.save(OUTPUT_FILE)
    print(f"Report saved to {OUTPUT_FILE}")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def main():
    print("Loading OTTO rate card...")
    otto_rates, name_to_id = load_otto_rates()
    print(f"  {len(otto_rates)} employees with rates")

    print("Loading kloklijst (timesheets)...")
    kloklijst = load_kloklijst()
    print(f"  {len(kloklijst)} employees in timesheets")

    print("Loading invoice...")
    invoice_lines, invoice_agg = load_invoice()
    print(f"  {len(invoice_lines)} invoice lines, {len(invoice_agg)} employees")

    print("Comparing hours...")
    hour_disc = compare_hours(invoice_agg, kloklijst)
    print(f"  {len(hour_disc)} hour discrepancies found")

    print("Comparing rates...")
    rate_disc = compare_rates(invoice_lines, otto_rates, name_to_id)
    print(f"  {len(rate_disc)} rate discrepancies found")

    print("Building employee summary...")
    employee_summary = build_employee_summary(invoice_lines, kloklijst, otto_rates, name_to_id)

    print("Writing report...")
    write_report(hour_disc, rate_disc, employee_summary)

    # Print quick summary
    total_invoiced = sum(l["subtotal"] for l in invoice_lines)
    total_hours_inv = sum(l["hours"] for l in invoice_lines)
    print(f"\n=== SUMMARY ===")
    print(f"Total invoiced: €{total_invoiced:,.2f}")
    print(f"Total hours invoiced: {total_hours_inv:,.1f}")
    print(f"Hour discrepancies: {len(hour_disc)}")
    print(f"Rate discrepancies: {len(rate_disc)}")


if __name__ == "__main__":
    main()
