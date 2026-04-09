import io
import unittest

try:
    from openpyxl import Workbook
except ImportError:  # pragma: no cover - environment dependency
    Workbook = None

from wagegroup_rates import (
    _pick_best_rate_card_candidate,
    _build_histogram_rows,
    _extract_schaal_tarief,
    _rate_key_from_code_toeslag,
    _rate_key_from_header,
    _same_wagegroup,
    _wagegroup_in_candidates,
    parse_flex_wagegroup_rate_workbook,
    parse_otto_wagegroup_rate_workbook,
)


class WagegroupRatesTests(unittest.TestCase):
    class _Card:
        def __init__(self, schaal: str, tarief: str, rate_value: float):
            self.schaal = schaal
            self.tarief = tarief
            self.rate_value = rate_value

    def test_rate_key_from_header(self):
        self.assertEqual(_rate_key_from_header("100%"), "100")
        self.assertEqual(_rate_key_from_header("1.33"), "133")
        self.assertEqual(_rate_key_from_header("2"), "200")

    def test_extract_schaal_tarief(self):
        schaal, tarief = _extract_schaal_tarief("Productiemdw D4 / Fase C")
        self.assertEqual(schaal, "D4")
        self.assertEqual(tarief, "C")

    def test_rate_key_from_code_toeslag(self):
        self.assertEqual(_rate_key_from_code_toeslag("T133 Dag"), "133")
        self.assertEqual(_rate_key_from_code_toeslag("OW200 Dag"), "200")
        self.assertEqual(_rate_key_from_code_toeslag("Norm uren Dag"), "100")

    def test_histogram_rows(self):
        rows = _build_histogram_rows([0.02, 0.4, 0.7, 1.2, 6.0])
        counts = {row["bucket"]: row["count"] for row in rows}
        self.assertEqual(counts["0.00-0.10"], 1)
        self.assertEqual(counts["0.25-0.50"], 1)
        self.assertEqual(counts["0.50-1.00"], 1)
        self.assertEqual(counts["1.00-2.00"], 1)
        self.assertEqual(counts["5.00+"], 1)

    def test_pick_best_rate_card_candidate_within_tolerance(self):
        candidates = [
            self._Card("A1", "A", 30.00),
            self._Card("B1", "B", 30.75),
            self._Card("C1", "C", 29.10),
        ]
        picked = _pick_best_rate_card_candidate(
            invoice_rate=30.6,
            candidates=candidates,
            tolerance_eur=1.0,
        )
        self.assertIsNotNone(picked)
        self.assertEqual(picked.schaal, "B1")
        self.assertEqual(picked.tarief, "B")

    def test_pick_best_rate_card_candidate_outside_tolerance(self):
        candidates = [self._Card("A1", "A", 30.0)]
        picked = _pick_best_rate_card_candidate(
            invoice_rate=32.5,
            candidates=candidates,
            tolerance_eur=1.0,
        )
        self.assertIsNone(picked)

    def test_same_wagegroup(self):
        self.assertTrue(
            _same_wagegroup(
                schaal_left="A1",
                tarief_left="A",
                schaal_right="A1",
                tarief_right="A",
            )
        )
        self.assertFalse(
            _same_wagegroup(
                schaal_left="A1",
                tarief_left="A",
                schaal_right="B1",
                tarief_right="A",
            )
        )

    def test_wagegroup_in_candidates(self):
        candidates = [
            self._Card("A1", "A", 30.0),
            self._Card("B1", "B", 31.0),
        ]
        self.assertTrue(
            _wagegroup_in_candidates(
                schaal="A1",
                tarief="A",
                candidates=candidates,
            )
        )
        self.assertFalse(
            _wagegroup_in_candidates(
                schaal="C1",
                tarief="C",
                candidates=candidates,
            )
        )

    @unittest.skipIf(Workbook is None, "openpyxl not installed")
    def test_parse_otto_workbook(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "Blad1"
        ws["A1"] = "Personeelsnummer"
        ws["B1"] = "Voornaam"
        ws["C1"] = "Achternaam"
        ws["D1"] = "100%"
        ws["E1"] = "133%"
        ws["F1"] = "135%"
        ws["G1"] = "180%"
        ws["H1"] = "200%"
        ws["I1"] = "300%"
        ws["A2"] = "441359"
        ws["B2"] = "Jan"
        ws["C2"] = "Jansen"
        ws["D2"] = 10
        ws["E2"] = 13
        ws["F2"] = 14
        ws["G2"] = 18
        ws["H2"] = 20
        ws["I2"] = 30
        buf = io.BytesIO()
        wb.save(buf)

        parsed = parse_otto_wagegroup_rate_workbook(buf.getvalue(), "Padifood tarieven p.p..xlsx")
        self.assertEqual(len(parsed.person_rates), 6)
        self.assertTrue(any(r["rate_key"] == "300" for r in parsed.person_rates))
        self.assertTrue(any(r["name"] == "Jan Jansen" for r in parsed.person_rates))

    @unittest.skipIf(Workbook is None, "openpyxl not installed")
    def test_parse_flex_workbook(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "1 januari 2025"
        ws["A1"] = "Loonnummer"
        ws["B1"] = "Voornaam"
        ws["C1"] = "Achternaam"
        ws["D1"] = "100%"
        ws["E1"] = "133%"
        ws["F1"] = "135%"
        ws["G1"] = "180%"
        ws["H1"] = "200%"
        ws["I1"] = "300%"
        ws["A2"] = "99123"
        ws["B2"] = "Piet"
        ws["C2"] = "Peters"
        ws["D2"] = 20.0
        ws["E2"] = 26.6
        ws["F2"] = 27.0
        ws["G2"] = 36.0
        ws["H2"] = 40.0
        ws["I2"] = 60.0
        buf = io.BytesIO()
        wb.save(buf)

        parsed = parse_flex_wagegroup_rate_workbook(buf.getvalue(), "Lonen en tarieven 2025.xlsx")
        self.assertEqual(len(parsed.person_rates), 6)
        self.assertTrue(any(r["person_number"] == "99123" for r in parsed.person_rates))
        self.assertTrue(any(r["name"] == "Piet Peters" for r in parsed.person_rates))


if __name__ == "__main__":
    unittest.main()
