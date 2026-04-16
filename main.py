from contextlib import asynccontextmanager
import csv
import io
import os
from pathlib import Path
import tempfile
from datetime import date, datetime
from typing import List
import re

from fastapi import Depends, FastAPI, UploadFile, File, HTTPException, Form
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse
from fastapi.security import APIKeyHeader
from openpyxl import load_workbook
from pydantic import BaseModel
from sqlalchemy import delete
from sqlalchemy.ext.asyncio import AsyncSession, async_sessionmaker, create_async_engine

from config import Settings
from convert import convert_input
from models import PersonWagegroup
from otto_identifier_mapping import (
    build_otto_mapping_candidates,
    persist_otto_mapping_candidates,
    write_otto_mapping_csv,
)
from validation_wagegroups import (
    analyze_otto_wagegroups,
    backfill_wagegroups_from_csv,
    list_person_wagegroups,
    load_otto_identity_context,
    resolve_wagegroup_identity,
    upsert_person_wagegroup,
    verify_otto_wagegroup_coverage,
)
from convert_flex import extract_week_from_flex_pdfs
from convert_otto_pdf import extract_week_from_otto_pdfs
from loaders import load_file, load_flex_invoices, load_otto_pdf_invoices
from validation_hours import (
    format_validation_email_body,
    normalize_name,
    run_validation,
)
from wagegroup_rates import (
    parse_flex_wagegroup_rate_workbook,
    parse_otto_wagegroup_rate_workbook,
    persist_parsed_wagegroup_rates,
    write_rates_csvs,
)

settings = Settings()

API_KEY_HEADER = APIKeyHeader(name="X-API-Key", auto_error=False)

# Support postgresql:// URL by converting to postgresql+asyncpg://
database_url = settings.postgres_database_url
if database_url.startswith("postgresql://"):
    database_url = database_url.replace("postgresql://", "postgresql+asyncpg://", 1)

engine = create_async_engine(database_url, echo=False)
async_session = async_sessionmaker(engine, class_=AsyncSession, expire_on_commit=False)


@asynccontextmanager
async def lifespan(app: FastAPI):
    yield
    await engine.dispose()


app = FastAPI(lifespan=lifespan)
BASE_DIR = Path(__file__).resolve().parent
OUTPUT_DIR = (BASE_DIR / "output").resolve()
WAGEGROUPS_CSV_PATH = (BASE_DIR / "person_wagegroups.csv").resolve()
WAGEGROUPS_HEADERS = ["id", "name", "wagegroup"]
VERIFIED_NAME_PAIRS_CSV_PATH = (BASE_DIR / "verified_name_pairs.csv").resolve()
VERIFIED_NAME_PAIRS_HEADERS = ["kloklijst_name", "factuur_name", "same_person"]


app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=False,
    allow_methods=["*"],
    allow_headers=["*"],
)


async def verify_api_key(api_key: str = Depends(API_KEY_HEADER)):
    if not api_key or api_key != settings.backend_api_key:
        raise HTTPException(
            status_code=401, detail="Ontbrekende of ongeldige API-sleutel"
        )


async def get_db() -> AsyncSession:
    async with async_session() as session:
        yield session


def _normalize_person_name(name: str) -> str:
    return " ".join(name.strip().lower().split())


def _normalize_rate_agency(agency: str) -> str:
    provider = (agency or "").strip().lower()
    if provider in {"otto", "flexspecialisten"}:
        return provider
    raise HTTPException(
        status_code=400, detail="agency moet 'otto' of 'flexspecialisten' zijn."
    )


def _normalize_header_cell(value: str) -> str:
    normalized = value.replace("\ufeff", "")
    normalized = re.sub(r"[\s,;:_-]+", " ", normalized.strip().lower())
    return " ".join(normalized.split())


def _is_bulk_header_row(name: str, wagegroup: str) -> bool:
    normalized_name = _normalize_header_cell(name)
    normalized_wagegroup = _normalize_header_cell(wagegroup)

    name_headers = {
        "name",
        "naam",
        "achternaam voornaam",
        "voornaam achternaam",
    }
    wagegroup_headers = {
        "wagegroup",
        "wage group",
        "loon groep",
        "loongroep",
    }

    if normalized_name not in name_headers:
        return False
    if not normalized_wagegroup:
        return True
    return normalized_wagegroup in wagegroup_headers


def _detect_wagegroup_columns(ws) -> tuple[int, int, int] | None:
    """
    Detect (name_col_idx, wagegroup_col_idx, header_row_idx) from sheet headers.
    Supports both legacy headers and OTTO-style headers.
    """
    name_headers = {
        "name",
        "naam",
        "achternaam voornaam",
        "voornaam achternaam",
    }
    wagegroup_headers = {
        "wagegroup",
        "wage group",
        "loon groep",
        "loongroep",
    }

    for row_idx, row in enumerate(ws.iter_rows(values_only=True, max_row=60), start=1):
        normalized_cells: list[str] = []
        for cell in row:
            if cell is None:
                normalized_cells.append("")
                continue
            normalized_cells.append(_normalize_header_cell(str(cell)))

        if not any(normalized_cells):
            continue

        name_idx = next(
            (
                idx
                for idx, value in enumerate(normalized_cells)
                if value in name_headers
            ),
            None,
        )
        wagegroup_idx = next(
            (
                idx
                for idx, value in enumerate(normalized_cells)
                if value in wagegroup_headers
            ),
            None,
        )
        if name_idx is not None and wagegroup_idx is not None:
            return name_idx, wagegroup_idx, row_idx

    return None


def _clean_person_fields(name: str, wagegroup: str) -> tuple[str, str]:
    cleaned_name = name.strip()
    cleaned_wagegroup = wagegroup.strip()
    if not cleaned_name or not cleaned_wagegroup:
        raise HTTPException(
            status_code=400,
            detail="Velden 'name' en 'wagegroup' zijn verplicht.",
        )
    return cleaned_name, cleaned_wagegroup


def _normalize_name_for_validation(name: str) -> str:
    return normalize_name(name)


def _ensure_verified_name_pairs_csv() -> None:
    if VERIFIED_NAME_PAIRS_CSV_PATH.exists():
        return
    with open(VERIFIED_NAME_PAIRS_CSV_PATH, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=VERIFIED_NAME_PAIRS_HEADERS)
        writer.writeheader()


def _read_verified_name_pairs() -> list[dict[str, str | bool]]:
    _ensure_verified_name_pairs_csv()
    with open(VERIFIED_NAME_PAIRS_CSV_PATH, newline="", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        missing_headers = [
            header
            for header in VERIFIED_NAME_PAIRS_HEADERS
            if header not in (reader.fieldnames or [])
        ]
        if missing_headers:
            raise HTTPException(
                status_code=500,
                detail=(
                    "CSV-indeling voor geverifieerde naamparen is ongeldig. "
                    f"Ontbrekende kolommen: {', '.join(missing_headers)}"
                ),
            )

        rows: list[dict[str, str | bool]] = []
        for row in reader:
            kloklijst_name = str(row.get("kloklijst_name", "")).strip()
            factuur_name = str(row.get("factuur_name", "")).strip()
            same_person_raw = str(row.get("same_person", "")).strip().lower()
            if not kloklijst_name or not factuur_name:
                continue
            same_person = same_person_raw in {"1", "true", "yes", "ja"}
            rows.append(
                {
                    "kloklijst_name": kloklijst_name,
                    "factuur_name": factuur_name,
                    "same_person": same_person,
                }
            )
        return rows


def _write_verified_name_pairs(rows: list[dict[str, str | bool]]) -> None:
    _ensure_verified_name_pairs_csv()
    fd, temp_path = tempfile.mkstemp(
        prefix="verified_name_pairs_",
        suffix=".csv",
        dir=str(BASE_DIR),
    )
    os.close(fd)

    try:
        with open(temp_path, "w", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(f, fieldnames=VERIFIED_NAME_PAIRS_HEADERS)
            writer.writeheader()
            for row in rows:
                writer.writerow(
                    {
                        "kloklijst_name": str(row.get("kloklijst_name", "")).strip(),
                        "factuur_name": str(row.get("factuur_name", "")).strip(),
                        "same_person": (
                            "true" if bool(row.get("same_person")) else "false"
                        ),
                    }
                )
        os.replace(temp_path, VERIFIED_NAME_PAIRS_CSV_PATH)
    finally:
        if os.path.exists(temp_path):
            os.remove(temp_path)


def _upsert_verified_name_pair(
    rows: list[dict[str, str | bool]],
    kloklijst_name: str,
    factuur_name: str,
    same_person: bool,
) -> None:
    target_kloklijst = _normalize_name_for_validation(kloklijst_name)
    target_factuur = _normalize_name_for_validation(factuur_name)

    for row in rows:
        existing_kloklijst = _normalize_name_for_validation(
            str(row.get("kloklijst_name", ""))
        )
        existing_factuur = _normalize_name_for_validation(
            str(row.get("factuur_name", ""))
        )
        if (
            existing_kloklijst == target_kloklijst
            and existing_factuur == target_factuur
        ):
            row["kloklijst_name"] = kloklijst_name
            row["factuur_name"] = factuur_name
            row["same_person"] = same_person
            return

    rows.append(
        {
            "kloklijst_name": kloklijst_name,
            "factuur_name": factuur_name,
            "same_person": same_person,
        }
    )


def _decision_pairs_for_validation(
    rows: list[dict[str, str | bool]],
) -> tuple[list[tuple[str, str]], list[tuple[str, str]]]:
    same_pairs: list[tuple[str, str]] = []
    diff_pairs: list[tuple[str, str]] = []
    for row in rows:
        kloklijst_name = str(row.get("kloklijst_name", "")).strip()
        factuur_name = str(row.get("factuur_name", "")).strip()
        if not kloklijst_name or not factuur_name:
            continue
        if bool(row.get("same_person")):
            same_pairs.append((kloklijst_name, factuur_name))
        else:
            diff_pairs.append((kloklijst_name, factuur_name))
    return same_pairs, diff_pairs


def _validate_week(week: str) -> str:
    cleaned_week = week.strip()
    if not re.fullmatch(r"\d{6}", cleaned_week):
        raise HTTPException(
            status_code=400,
            detail="Week moet in formaat YYYYww zijn, bijvoorbeeld 202551.",
        )
    return cleaned_week


def _ensure_wagegroups_csv() -> None:
    if WAGEGROUPS_CSV_PATH.exists():
        return
    with open(WAGEGROUPS_CSV_PATH, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=WAGEGROUPS_HEADERS)
        writer.writeheader()


def _read_wagegroups() -> list[dict[str, str]]:
    _ensure_wagegroups_csv()
    with open(WAGEGROUPS_CSV_PATH, newline="", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        missing_headers = [
            header
            for header in WAGEGROUPS_HEADERS
            if header not in (reader.fieldnames or [])
        ]
        if missing_headers:
            raise HTTPException(
                status_code=500,
                detail=(
                    "CSV-indeling voor wagegroups is ongeldig. "
                    f"Ontbrekende kolommen: {', '.join(missing_headers)}"
                ),
            )

        rows: list[dict[str, str]] = []
        for row in reader:
            id_value = str(row.get("id", "")).strip()
            name = str(row.get("name", "")).strip()
            wagegroup = str(row.get("wagegroup", "")).strip()

            if not id_value and not name and not wagegroup:
                continue

            rows.append(
                {
                    "id": id_value,
                    "name": name,
                    "wagegroup": wagegroup,
                }
            )
        return rows


def _write_wagegroups(rows: list[dict[str, str]]) -> None:
    _ensure_wagegroups_csv()
    fd, temp_path = tempfile.mkstemp(
        prefix="person_wagegroups_",
        suffix=".csv",
        dir=str(BASE_DIR),
    )
    os.close(fd)

    try:
        with open(temp_path, "w", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(f, fieldnames=WAGEGROUPS_HEADERS)
            writer.writeheader()
            for row in rows:
                writer.writerow(
                    {
                        "id": str(row.get("id", "")).strip(),
                        "name": str(row.get("name", "")).strip(),
                        "wagegroup": str(row.get("wagegroup", "")).strip(),
                    }
                )
        os.replace(temp_path, WAGEGROUPS_CSV_PATH)
    finally:
        if os.path.exists(temp_path):
            os.remove(temp_path)


def _next_person_id(rows: list[dict[str, str]]) -> int:
    max_id = 0
    for row in rows:
        try:
            max_id = max(max_id, int(str(row.get("id", "")).strip()))
        except ValueError:
            continue
    return max_id + 1


def _upsert_person_wagegroup(
    rows: list[dict[str, str]], name: str, wagegroup: str
) -> str:
    normalized_name = _normalize_person_name(name)
    for row in rows:
        if _normalize_person_name(str(row.get("name", ""))) == normalized_name:
            row["name"] = name
            row["wagegroup"] = wagegroup
            return "updated"

    rows.append(
        {
            "id": str(_next_person_id(rows)),
            "name": name,
            "wagegroup": wagegroup,
        }
    )
    return "inserted"


@app.get("/health")
def health():
    return {"status": "ok"}


class DownloadRequest(BaseModel):
    fileName: str


class WagePersonUpdateRequest(BaseModel):
    provider: str = "otto"
    personNumber: str | None = None
    kloklijstLoonnummer: str | None = None
    name: str
    wagegroup: str
    verified: bool = True
    sourceWeek: int | None = None


class WagegroupBackfillRequest(BaseModel):
    week: str
    provider: str = "otto"
    csvPath: str | None = None


class OttoWagegroupAnalysisRequest(BaseModel):
    week: str
    includeMismatches: bool = True
    maxItems: int = 500


class OttoWagegroupCoverageRequest(BaseModel):
    week: str


class NamePairDecision(BaseModel):
    kloklijstName: str
    factuurName: str
    samePerson: bool


class VerifyNamePairsRequest(BaseModel):
    week: str
    agency: str = "otto"
    decisions: list[NamePairDecision]


class OttoIdentifierMappingBuildRequest(BaseModel):
    week: str
    persist: bool = False
    requireFullCoverage: bool = False
    writeCsv: bool = True
    includeCandidates: bool = False


@app.post("/download")
async def download(payload: DownloadRequest):
    requested = payload.fileName.strip()
    if not requested:
        raise HTTPException(status_code=400, detail="fileName is required.")

    candidate_path = Path(requested)
    if not candidate_path.is_absolute():
        candidate_path = (BASE_DIR / candidate_path).resolve()
    else:
        candidate_path = candidate_path.resolve()

    try:
        candidate_path.relative_to(OUTPUT_DIR)
    except ValueError as e:
        raise HTTPException(
            status_code=400,
            detail="Ongeldig bestandspad.",
        ) from e

    if not candidate_path.exists() or not candidate_path.is_file():
        raise HTTPException(status_code=404, detail="Bestand niet gevonden.")

    return FileResponse(
        path=str(candidate_path),
        filename=candidate_path.name,
        media_type="application/octet-stream",
    )


@app.post("/otto_identifier_mapping/build")
async def build_otto_identifier_mapping(
    payload: OttoIdentifierMappingBuildRequest,
    db: AsyncSession = Depends(get_db),
    _: None = Depends(verify_api_key),
):
    week = _validate_week(payload.week)
    week_int = int(week)
    analysis = await build_otto_mapping_candidates(week_int, db)
    stats = analysis["stats"]

    csv_path = None
    if payload.writeCsv:
        csv_path = write_otto_mapping_csv(analysis["candidates"])

    if payload.requireFullCoverage and not bool(stats.get("coverage100")):
        raise HTTPException(
            status_code=400,
            detail=(
                "Coverage is niet 100% voor Otto mapping. "
                "Controleer unresolved people en uniqueness conflicts."
            ),
        )

    persist_result = {"insertedMappings": 0}
    if payload.persist:
        persist_result = await persist_otto_mapping_candidates(
            week=week_int,
            candidates=analysis["candidates"],
            db=db,
        )

    response = {
        "status": "ok",
        "stats": stats,
        "csvBackupPath": csv_path,
        "uniquenessConflicts": analysis["uniquenessConflicts"],
        "persistResult": persist_result,
    }
    if payload.includeCandidates:
        response["candidates"] = analysis["candidates"]
    return response


@app.get("/wagegroups")
async def get_wagegroups(
    provider: str | None = None,
    db: AsyncSession = Depends(get_db),
    _: None = Depends(verify_api_key),
):
    rows = await list_person_wagegroups(db=db, provider=provider)
    return [
        {
            "id": row.id,
            "provider": row.provider,
            "personNumber": row.person_number,
            "kloklijstLoonnummer": row.kloklijst_loonnummer,
            "name": row.name,
            "wagegroup": row.wagegroup,
            "verified": row.verified,
            "sourceWeek": row.source_week,
        }
        for row in rows
    ]


async def _clear_wagegroups_db(
    db: AsyncSession,
    provider: str | None = None,
) -> dict[str, object]:
    provider_normalized = provider.strip().lower() if provider else None
    stmt = delete(PersonWagegroup)
    if provider_normalized:
        stmt = stmt.where(PersonWagegroup.provider == provider_normalized)
    result = await db.execute(stmt)
    await db.commit()
    return {
        "status": "ok",
        "provider": provider_normalized or "all",
        "deleted": int(result.rowcount or 0),
    }


@app.delete("/wagegroups")
async def delete_wagegroups(
    provider: str | None = None,
    db: AsyncSession = Depends(get_db),
    _: None = Depends(verify_api_key),
):
    return await _clear_wagegroups_db(db=db, provider=provider)


@app.post("/wagegroups/clear")
async def clear_wagegroups(
    provider: str | None = None,
    db: AsyncSession = Depends(get_db),
    _: None = Depends(verify_api_key),
):
    return await _clear_wagegroups_db(db=db, provider=provider)


@app.post("/update_wage_person")
async def update_wage_person(
    payload: WagePersonUpdateRequest,
    db: AsyncSession = Depends(get_db),
    _: None = Depends(verify_api_key),
):
    name, wagegroup = _clean_person_fields(payload.name, payload.wagegroup)
    provider = payload.provider.strip().lower()
    if not provider:
        raise HTTPException(status_code=400, detail="provider is verplicht.")
    identity = await resolve_wagegroup_identity(
        db,
        provider=provider,
        raw_name=name,
        supplied_person_number=(
            payload.personNumber.strip() if payload.personNumber else None
        ),
    )
    person_number = str(identity["person_number"])
    resolved_name = str(identity["name"])
    resolved_kloklijst_loonnummer = (
        payload.kloklijstLoonnummer.strip()
        if payload.kloklijstLoonnummer
        else (
            str(identity["kloklijst_loonnummer"])
            if identity.get("kloklijst_loonnummer")
            else None
        )
    )
    resolved_verified = payload.verified or bool(identity.get("verified"))

    action = await upsert_person_wagegroup(
        db,
        provider=provider,
        person_number=person_number,
        name=resolved_name,
        wagegroup=wagegroup,
        kloklijst_loonnummer=resolved_kloklijst_loonnummer,
        verified=resolved_verified,
        source_week=payload.sourceWeek,
    )
    return {
        "status": "ok",
        "action": action,
        "personNumber": person_number,
        "identityMatchMethod": identity.get("match_method"),
    }


@app.post("/update_wages")
async def update_wages(
    file: UploadFile = File(...),
    provider: str = "otto",
    db: AsyncSession = Depends(get_db),
    _: None = Depends(verify_api_key),
):
    if not file.filename:
        raise HTTPException(status_code=400, detail="Geen bestandsnaam ontvangen.")
    if not file.filename.lower().endswith(".xlsx"):
        raise HTTPException(
            status_code=400, detail="Alleen .xlsx-bestanden worden ondersteund."
        )

    content = await file.read()
    if not content:
        raise HTTPException(status_code=400, detail="Leeg bestand ontvangen.")

    try:
        wb = load_workbook(io.BytesIO(content), data_only=True)
    except Exception as e:
        raise HTTPException(
            status_code=400, detail="Kon Excelbestand niet lezen."
        ) from e

    ws = wb.active
    detected_columns = _detect_wagegroup_columns(ws)
    if detected_columns is None:
        raise HTTPException(
            status_code=400,
            detail=(
                "Kon kolommen voor naam en loongroep niet herkennen in het Excelbestand. "
                "Verwachte headers zijn bijvoorbeeld: "
                "'Achternaam voornaam' + 'Loongroep' of 'Name/Naam' + 'Wagegroup'."
            ),
        )

    name_col_idx, wagegroup_col_idx, header_row_idx = detected_columns
    provider_normalized = provider.strip().lower()
    if not provider_normalized:
        raise HTTPException(status_code=400, detail="provider is verplicht.")

    processed = 0
    inserted = 0
    updated = 0
    skipped = 0
    identity_resolved = 0
    embedded_numbers_detected = 0
    otto_context = (
        await load_otto_identity_context(db) if provider_normalized == "otto" else None
    )

    for row_idx, excel_row in enumerate(ws.iter_rows(values_only=True), start=1):
        if row_idx <= header_row_idx:
            continue

        raw_name = excel_row[name_col_idx] if len(excel_row) > name_col_idx else None
        raw_wagegroup = (
            excel_row[wagegroup_col_idx] if len(excel_row) > wagegroup_col_idx else None
        )

        name = str(raw_name).strip() if raw_name is not None else ""
        wagegroup = str(raw_wagegroup).strip() if raw_wagegroup is not None else ""

        if _is_bulk_header_row(name, wagegroup):
            skipped += 1
            continue

        if not name and not wagegroup:
            continue

        if not name or not wagegroup:
            skipped += 1
            continue

        identity = await resolve_wagegroup_identity(
            db,
            provider=provider_normalized,
            raw_name=name,
            otto_context=otto_context,
        )
        person_number = str(identity["person_number"])
        resolved_name = str(identity["name"])
        resolved_kloklijst_loonnummer = (
            str(identity["kloklijst_loonnummer"])
            if identity.get("kloklijst_loonnummer")
            else None
        )
        if identity.get("verified"):
            identity_resolved += 1
        if identity.get("kloklijst_loonnummer"):
            embedded_numbers_detected += 1

        processed += 1
        action = await upsert_person_wagegroup(
            db,
            provider=provider_normalized,
            person_number=person_number,
            name=resolved_name,
            wagegroup=wagegroup,
            kloklijst_loonnummer=resolved_kloklijst_loonnummer,
            verified=bool(identity.get("verified")),
        )
        if action == "inserted":
            inserted += 1
        else:
            updated += 1

    return {
        "status": "ok",
        "detectedColumns": {
            "nameColumnIndex": name_col_idx,
            "wagegroupColumnIndex": wagegroup_col_idx,
            "headerRowIndex": header_row_idx,
        },
        "processed": processed,
        "inserted": inserted,
        "updated": updated,
        "skipped": skipped,
        "identityResolved": identity_resolved,
        "embeddedNumbersDetected": embedded_numbers_detected,
    }


@app.post("/upload_wagegroups_rate")
async def upload_wagegroups_rate(
    file: UploadFile = File(...),
    agency: str = Form("otto"),
    db: AsyncSession = Depends(get_db),
    _: None = Depends(verify_api_key),
):
    if not file.filename:
        raise HTTPException(status_code=400, detail="Geen bestandsnaam ontvangen.")
    if not file.filename.lower().endswith(".xlsx"):
        raise HTTPException(
            status_code=400, detail="Alleen .xlsx-bestanden worden ondersteund."
        )

    provider = _normalize_rate_agency(agency)

    content = await file.read()
    if not content:
        raise HTTPException(status_code=400, detail="Leeg bestand ontvangen.")

    try:
        if provider == "otto":
            parsed = parse_otto_wagegroup_rate_workbook(content, file.filename)
        else:
            parsed = parse_flex_wagegroup_rate_workbook(content, file.filename)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e)) from e
    except Exception as e:
        raise HTTPException(
            status_code=400, detail=f"Kon tarievenbestand niet verwerken: {e}"
        ) from e

    db_stats = await persist_parsed_wagegroup_rates(
        db, provider=provider, parsed=parsed
    )
    csv_paths = write_rates_csvs(
        provider=provider,
        parsed=parsed,
        output_dir=str(OUTPUT_DIR),
    )
    return {
        "status": "ok",
        "agency": provider,
        "personRates": db_stats["personRates"],
        "rateCardRows": db_stats["rateCardRows"],
        **csv_paths,
    }


@app.post("/wagegroups/backfill_from_csv")
async def backfill_wagegroups(
    payload: WagegroupBackfillRequest,
    db: AsyncSession = Depends(get_db),
    _: None = Depends(verify_api_key),
):
    week = _validate_week(payload.week)
    csv_path = (
        Path(payload.csvPath).resolve() if payload.csvPath else WAGEGROUPS_CSV_PATH
    )
    try:
        result = await backfill_wagegroups_from_csv(
            week=int(week),
            db=db,
            csv_path=csv_path,
            provider=payload.provider.strip().lower(),
        )
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e)) from e
    return result


@app.post("/otto_wagegroups/analyze")
async def analyze_otto_wagegroups_endpoint(
    payload: OttoWagegroupAnalysisRequest,
    db: AsyncSession = Depends(get_db),
    _: None = Depends(verify_api_key),
):
    week = _validate_week(payload.week)
    return await analyze_otto_wagegroups(
        week=int(week),
        db=db,
        include_mismatches=payload.includeMismatches,
        max_items=max(1, payload.maxItems),
    )


@app.post("/otto_wagegroups/verify_coverage")
async def verify_otto_wagegroups_coverage_endpoint(
    payload: OttoWagegroupCoverageRequest,
    db: AsyncSession = Depends(get_db),
    _: None = Depends(verify_api_key),
):
    week = _validate_week(payload.week)
    return await verify_otto_wagegroup_coverage(week=int(week), db=db)


@app.post("/upload")
async def upload(
    files: List[UploadFile] = File(...),
    db: AsyncSession = Depends(get_db),
    _: None = Depends(verify_api_key),
):
    input_dir = "input"
    os.makedirs(input_dir, exist_ok=True)

    if len(files) < 2:
        raise HTTPException(
            status_code=400,
            detail="Minimaal 2 bestanden vereist.",
        )

    # ------------------------------------------------------------------ helpers

    def extract_week_from_filename(filename: str) -> str | None:
        match = re.match(r"^\s*(\d{6})\b", filename)
        return match.group(1) if match else None

    def parse_excel_date(value) -> date | None:
        if isinstance(value, datetime):
            return value.date()
        if isinstance(value, date):
            return value
        if value is None:
            return None
        text = str(value).strip()
        if not text:
            return None
        text = text.split(" ")[0]
        for fmt in ["%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y", "%Y/%m/%d"]:
            try:
                return datetime.strptime(text, fmt).date()
            except ValueError:
                continue
        return None

    def extract_week_from_factuur(content: bytes) -> str | None:
        wb = load_workbook(io.BytesIO(content), data_only=True)
        if "Export Factuur" not in wb.sheetnames:
            return None
        ws = wb["Export Factuur"]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return None
        header = [str(c).strip() if c is not None else "" for c in rows[0]]
        if "Datum" not in header:
            return None
        datum_idx = header.index("Datum")
        for row in rows[1:]:
            if datum_idx >= len(row):
                continue
            parsed = parse_excel_date(row[datum_idx])
            if parsed is None:
                continue
            iso = parsed.isocalendar()
            return f"{iso[0]}{iso[1]:02d}"
        return None

    def build_prefixed_filename(original_filename: str, week: str) -> str:
        fname = os.path.basename(original_filename)
        stem = re.sub(r"\.xlsx$", "", fname, flags=re.IGNORECASE)
        stem = re.sub(r"^\s*\d{6}\b[\s,._-]*", "", stem)
        stem = stem.replace(",", " ")
        stem = " ".join(stem.split()).strip()
        if not stem:
            raise HTTPException(
                status_code=400,
                detail=f"Kon geen geldige bestandsnaam afleiden uit: {original_filename}",
            )
        return f"{week} {stem}.xlsx"

    # ------------------------------------------------------------------ read + classify

    uploaded = []
    for file in files:
        if not file.filename:
            raise HTTPException(
                status_code=400, detail="Geüpload bestand heeft geen bestandsnaam."
            )
        lower = file.filename.lower()
        if not (lower.endswith(".xlsx") or lower.endswith(".pdf")):
            raise HTTPException(
                status_code=400,
                detail=f"Alleen .xlsx en .pdf bestanden worden ondersteund: {file.filename}",
            )
        content = await file.read()
        uploaded.append({"filename": file.filename, "content": content})

    otto_kloklijst: dict | None = None
    flex_kloklijst: dict | None = None
    otto_factuur: dict | None = None
    flex_pdfs: list[dict] = []
    otto_pdfs: list[dict] = []

    for item in uploaded:
        lower = item["filename"].lower()
        is_pdf = lower.endswith(".pdf")
        is_kloklijst = "kloklijst" in lower

        is_flex = "flex" in lower
        is_otto = "otto" in lower

        if is_pdf:
            if is_otto:
                otto_pdfs.append(item)
            elif is_flex:
                flex_pdfs.append(item)
            else:
                raise HTTPException(
                    status_code=400,
                    detail=f"Kan agency niet bepalen voor PDF: {item['filename']}. Bestandsnaam moet 'Otto' of 'Flex' bevatten.",
                )
        elif is_kloklijst:
            if is_otto:
                if otto_kloklijst is not None:
                    raise HTTPException(
                        status_code=400,
                        detail="Meerdere OTTO kloklijstbestanden gevonden.",
                    )
                otto_kloklijst = item
            elif is_flex:
                if flex_kloklijst is not None:
                    raise HTTPException(
                        status_code=400,
                        detail="Meerdere Flexspecialisten kloklijstbestanden gevonden.",
                    )
                flex_kloklijst = item
            else:
                raise HTTPException(
                    status_code=400,
                    detail=f"Kan agency niet bepalen voor kloklijst: {item['filename']}. Bestandsnaam moet 'Otto' of 'Flex' bevatten.",
                )
        elif "factuur" in lower or "specificatie" in lower:
            if is_flex:
                # Flex XLSX invoice — not currently processed but don't block
                pass
            else:
                if otto_factuur is not None:
                    raise HTTPException(
                        status_code=400,
                        detail="Meerdere OTTO factuurbestanden gevonden.",
                    )
                otto_factuur = item
        elif "tarief" in lower or "lonen" in lower:
            # Tarievensheet uploaded as part of batch — ignore here, handled by /upload_tarievensheet
            pass
        else:
            raise HTTPException(
                status_code=400,
                detail=f"Onbekend bestandstype: {item['filename']}",
            )

    has_otto_xlsx = otto_kloklijst is not None and otto_factuur is not None
    has_otto_pdf = otto_kloklijst is not None and len(otto_pdfs) > 0
    has_otto = has_otto_xlsx or has_otto_pdf
    has_flex = flex_kloklijst is not None and len(flex_pdfs) > 0

    if not has_otto and not has_flex:
        raise HTTPException(
            status_code=400,
            detail=(
                "Geen volledig paar gevonden. Geef een OTTO kloklijst+factuurbestand (xlsx of pdf) "
                "en/of een Flexspecialisten kloklijst+PDF."
            ),
        )
    if otto_factuur is not None and len(otto_pdfs) > 0:
        raise HTTPException(
            status_code=400,
            detail="Zowel een OTTO factuur xlsx als OTTO PDF gevonden. Geef één van beide.",
        )
    if otto_kloklijst and not otto_factuur and not otto_pdfs:
        raise HTTPException(
            status_code=400,
            detail="OTTO kloklijst gevonden maar geen OTTO factuur (xlsx of pdf).",
        )
    if otto_factuur and not otto_kloklijst:
        raise HTTPException(
            status_code=400,
            detail="OTTO factuurbestand gevonden maar geen OTTO kloklijst.",
        )
    if flex_kloklijst and not flex_pdfs:
        raise HTTPException(
            status_code=400,
            detail="Flexspecialisten kloklijst gevonden maar geen PDF factuur.",
        )
    if flex_pdfs and not flex_kloklijst:
        raise HTTPException(
            status_code=400,
            detail="Flexspecialisten PDF gevonden maar geen Flexspecialisten kloklijst.",
        )

    # ------------------------------------------------------------------ week extraction + cross-validation

    week: str | None = None

    if has_otto:
        otto_week_kl = extract_week_from_filename(otto_kloklijst["filename"])  # type: ignore[index]
        if not otto_week_kl:
            raise HTTPException(
                status_code=400,
                detail="Weeknummer (YYYYww) niet gevonden in OTTO kloklijst bestandsnaam.",
            )
        if has_otto_xlsx:
            otto_week_fa = extract_week_from_factuur(otto_factuur["content"])  # type: ignore[index]
            if not otto_week_fa:
                print(
                    "WARNING: Weeknummer niet te bepalen uit kolom Datum in OTTO factuur — week uit kloklijst gebruikt"
                )
            elif otto_week_kl != otto_week_fa:
                print(
                    f"WARNING: OTTO weeknummers komen niet overeen: kloklijst={otto_week_kl}, factuur={otto_week_fa} — doorgaan met kloklijst week"
                )
        elif has_otto_pdf:
            otto_week_pdf_raw = extract_week_from_otto_pdfs(
                [p["content"] for p in otto_pdfs]
            )
            if otto_week_pdf_raw is None:
                raise HTTPException(
                    status_code=400,
                    detail="Weeknummer niet te bepalen uit de OTTO PDF.",
                )
            otto_week_pdf = str(otto_week_pdf_raw)
            if otto_week_kl != otto_week_pdf:
                print(
                    f"WARNING: OTTO weeknummers komen niet overeen: kloklijst={otto_week_kl}, PDF={otto_week_pdf} — doorgaan met kloklijst week"
                )
        week = otto_week_kl

    if has_flex:
        flex_week_kl = extract_week_from_filename(flex_kloklijst["filename"])  # type: ignore[index]
        if not flex_week_kl:
            raise HTTPException(
                status_code=400,
                detail="Weeknummer (YYYYww) niet gevonden in Flexspecialisten kloklijst bestandsnaam.",
            )
        flex_week_pdf_raw = extract_week_from_flex_pdfs(
            [p["content"] for p in flex_pdfs]
        )
        if flex_week_pdf_raw is None:
            raise HTTPException(
                status_code=400,
                detail="Weeknummer niet te bepalen uit de Flexspecialisten PDF.",
            )
        flex_week_pdf = str(flex_week_pdf_raw)
        if flex_week_kl != flex_week_pdf:
            print(
                f"WARNING: Flex weeknummers komen niet overeen: kloklijst={flex_week_kl}, PDF={flex_week_pdf} — doorgaan met kloklijst week"
            )
        if week is not None and week != flex_week_kl:
            print(
                f"WARNING: OTTO weeknummer ({week}) komt niet overeen met Flexspecialisten weeknummer ({flex_week_kl}) — doorgaan"
            )
        week = flex_week_kl

    assert week is not None

    # ------------------------------------------------------------------ load name-pair decisions

    verified_rows = _read_verified_name_pairs()
    confirmed_same_pairs, confirmed_diff_pairs = _decision_pairs_for_validation(
        verified_rows
    )

    # ------------------------------------------------------------------ process OTTO

    results: list[dict] = []

    if has_otto:
        kloklijst_name = build_prefixed_filename(otto_kloklijst["filename"], week)  # type: ignore[index]
        with open(os.path.join(input_dir, kloklijst_name), "wb") as f:
            f.write(otto_kloklijst["content"])  # type: ignore[index]

        try:
            if has_otto_xlsx:
                factuur_name = build_prefixed_filename(otto_factuur["filename"], week)  # type: ignore[index]
                with open(os.path.join(input_dir, factuur_name), "wb") as f:
                    f.write(otto_factuur["content"])  # type: ignore[index]
                created_files = convert_input(kloklijst_name, factuur_name)
            else:
                # PDF invoice: convert kloklijst only, save PDFs for audit trail
                for p in otto_pdfs:
                    pdf_name = f"{week} {os.path.basename(p['filename'])}"
                    with open(os.path.join(input_dir, pdf_name), "wb") as f:
                        f.write(p["content"])
                created_files = convert_input(kloklijst_name)

            for file_path in created_files:
                fname = os.path.basename(file_path)
                with open(file_path, "rb") as f:
                    csv_content = f.read()
                await load_file(fname, csv_content, db)

            if has_otto_pdf:
                await load_otto_pdf_invoices(
                    [(p["filename"], p["content"]) for p in otto_pdfs],
                    int(week),
                    db,
                )
            otto_result = await run_validation(
                week,
                db,
                agency="otto",
                confirmed_same_pairs=confirmed_same_pairs,
                confirmed_diff_pairs=confirmed_diff_pairs,
            )
        except FileNotFoundError as e:
            raise HTTPException(status_code=400, detail=str(e)) from e
        except ValueError as e:
            raise HTTPException(status_code=400, detail=str(e)) from e
        except HTTPException:
            raise
        except Exception as e:
            import traceback

            traceback.print_exc()
            raise HTTPException(
                status_code=500, detail=f"Fout bij OTTO verwerking: {e}"
            )

        if not os.path.exists(otto_result["outputFileWeek"]) or not os.path.exists(
            otto_result["outputFileDay"]
        ):
            raise HTTPException(
                status_code=400,
                detail="OTTO outputbestanden niet aangemaakt door validatie.",
            )

        results.append(
            {
                "agency": "otto",
                "emailBody": format_validation_email_body(otto_result),
                "outputFileWeek": otto_result["outputFileWeek"],
                "outputFileDay": otto_result["outputFileDay"],
                "rateOutputFile": otto_result.get("rateOutputFile"),
                "rateHistogramFile": otto_result.get("rateHistogramFile"),
                "wagegroupRateOutputFile": otto_result.get("wagegroupRateOutputFile"),
                "wagegroupRateSummary": otto_result.get("wagegroupRateAnalysis") or {},
                "similarPeople": otto_result.get("similarPeople", []),
            }
        )

    # ------------------------------------------------------------------ process Flex

    if has_flex:
        flex_kl_name = build_prefixed_filename(flex_kloklijst["filename"], week)  # type: ignore[index]
        with open(os.path.join(input_dir, flex_kl_name), "wb") as f:
            f.write(flex_kloklijst["content"])  # type: ignore[index]

        try:
            # Convert and load the Flex kloklijst only (no paired xlsx invoice)
            created_files = convert_input(flex_kl_name)
            for file_path in created_files:
                fname = os.path.basename(file_path)
                with open(file_path, "rb") as f:
                    csv_content = f.read()
                await load_file(fname, csv_content, db)

            # Write PDFs to input/ for audit trail
            for p in flex_pdfs:
                pdf_name = build_prefixed_filename(p["filename"], week)
                with open(os.path.join(input_dir, pdf_name), "wb") as f:
                    f.write(p["content"])

            # Load Flex PDF invoices (with correction merging)
            await load_flex_invoices(
                [(p["filename"], p["content"]) for p in flex_pdfs],
                int(week),
                db,
            )

            flex_result = await run_validation(
                week,
                db,
                agency="flexspecialisten",
                confirmed_same_pairs=confirmed_same_pairs,
                confirmed_diff_pairs=confirmed_diff_pairs,
            )
        except FileNotFoundError as e:
            raise HTTPException(status_code=400, detail=str(e)) from e
        except ValueError as e:
            raise HTTPException(status_code=400, detail=str(e)) from e
        except HTTPException:
            raise
        except Exception as e:
            import traceback

            traceback.print_exc()
            raise HTTPException(
                status_code=500, detail=f"Fout bij Flexspecialisten verwerking: {e}"
            )

        if not os.path.exists(flex_result["outputFileWeek"]) or not os.path.exists(
            flex_result["outputFileDay"]
        ):
            raise HTTPException(
                status_code=400,
                detail="Flexspecialisten outputbestanden niet aangemaakt door validatie.",
            )

        results.append(
            {
                "agency": "flexspecialisten",
                "emailBody": format_validation_email_body(flex_result),
                "outputFileWeek": flex_result["outputFileWeek"],
                "outputFileDay": flex_result["outputFileDay"],
                "similarPeople": flex_result.get("similarPeople", []),
            }
        )

    return {"results": results}


@app.post("/verify_name_pairs")
async def verify_name_pairs(
    payload: VerifyNamePairsRequest,
    db: AsyncSession = Depends(get_db),
    _: None = Depends(verify_api_key),
):
    week = _validate_week(payload.week)
    rows = _read_verified_name_pairs()

    for decision in payload.decisions:
        kloklijst_name = decision.kloklijstName.strip()
        factuur_name = decision.factuurName.strip()
        if not kloklijst_name or not factuur_name:
            raise HTTPException(
                status_code=400,
                detail="Elke beslissing moet zowel kloklijstName als factuurName bevatten.",
            )
        _upsert_verified_name_pair(
            rows=rows,
            kloklijst_name=kloklijst_name,
            factuur_name=factuur_name,
            same_person=decision.samePerson,
        )

    _write_verified_name_pairs(rows)

    confirmed_same_pairs, confirmed_diff_pairs = _decision_pairs_for_validation(rows)
    providers_response: dict[str, dict] = {}
    validation_errors: list[str] = []
    try:
        for response_key, agency, provider_label in (
            ("otto", "otto", "Otto Workforce"),
            ("flex", "flexspecialisten", "Flexspecialisten"),
        ):
            try:
                validation_result = await run_validation(
                    week,
                    db,
                    agency=agency,
                    confirmed_same_pairs=confirmed_same_pairs,
                    confirmed_diff_pairs=confirmed_diff_pairs,
                )
            except ValueError as e:
                if "Geen kloklijstregels gevonden" in str(e):
                    continue
                validation_errors.append(str(e))
                continue

            output_file_week = validation_result["outputFileWeek"]
            wagegroup_output_file = validation_result.get("wagegroupOutputFile")
            if not os.path.exists(output_file_week):
                raise HTTPException(
                    status_code=400,
                    detail=(
                        "Het verwachte week-outputbestand is niet gegenereerd door de "
                        f"validatie voor {provider_label}."
                    ),
                )
            if (
                response_key == "otto"
                and wagegroup_output_file
                and not os.path.exists(wagegroup_output_file)
            ):
                raise HTTPException(
                    status_code=400,
                    detail=(
                        "Het verwachte wagegroup-outputbestand is niet gegenereerd door de "
                        f"validatie voor {provider_label}."
                    ),
                )

            provider_result = {**validation_result, "providerLabel": provider_label}
            provider_response = {
                "emailBody": format_validation_email_body(provider_result),
                "outputFileWeek": output_file_week,
                "exactPersonMatchCount": validation_result.get(
                    "exactPersonMatchCount", 0
                ),
                "wagegroupRateSummary": (
                    validation_result.get("wagegroupRateAnalysis") or {}
                ),
                "wagegroupRateOutputFile": validation_result.get(
                    "wagegroupRateOutputFile"
                ),
            }
            if response_key == "otto":
                provider_response["wagegroupOutputFile"] = wagegroup_output_file
                provider_response["wagegroupSummary"] = (
                    validation_result.get("wagegroupAnalysis") or {}
                )
                provider_response["rateSummary"] = (
                    validation_result.get("rateAnalysis") or {}
                )
                provider_response["rateOutputFile"] = validation_result.get(
                    "rateOutputFile"
                )
                provider_response["rateHistogramFile"] = validation_result.get(
                    "rateHistogramFile"
                )
            providers_response[response_key] = provider_response
    except FileNotFoundError as e:
        raise HTTPException(status_code=400, detail=str(e)) from e
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail=f"Onverwachte fout tijdens het herberekenen van validatie: {e}",
        ) from e

    if not providers_response:
        if validation_errors:
            raise HTTPException(status_code=400, detail=validation_errors[0])
        raise HTTPException(
            status_code=400,
            detail="Geen validatieresultaten gevonden voor Otto of Flexspecialisten.",
        )
    return {"providers": providers_response}
