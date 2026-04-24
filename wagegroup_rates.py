from __future__ import annotations

import csv
import io
import os
import re
import json
import uuid
from collections import Counter, defaultdict
from dataclasses import asdict, dataclass
from typing import Any

try:
    from openpyxl import load_workbook
except ImportError:  # pragma: no cover - runtime dependency
    load_workbook = None
from sqlalchemy import delete, select
from sqlalchemy.ext.asyncio import AsyncSession

from models import InvoiceLine, PersonWagegroup, PersonWagegroupRate, WagegroupRateCard
from validation_wagegroups import normalize_person_name

def _norm_text(value: Any) -> str:
    if value is None:
        return ""
    return " ".join(str(value).strip().split())


def _to_float(value: Any) -> float | None:
    if value is None:
        return None
    text = str(value).strip().replace(",", ".")
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def _to_person_number(value: Any) -> str | None:
    text = _norm_text(value)
    if not text:
        return None
    if text.endswith(".0"):
        text = text[:-2]
    return text


def _rate_key_from_header(value: Any) -> str | None:
    text = _norm_text(value).lower()
    if not text:
        return None
    text = text.replace("%", "")
    if text in {"1", "1.0", "100"}:
        return "100"
    if text in {"1.3", "130"}:
        return "130"
    if text in {"1.33", "133"}:
        return "133"
    if text in {"1.35", "135"}:
        return "135"
    if text in {"1.4", "140"}:
        return "140"
    if text in {"1.8", "180"}:
        return "180"
    if text in {"2", "2.0", "200"}:
        return "200"
    if text in {"3", "3.0", "300"}:
        return "300"
    return None


def _rate_key_from_code_toeslag(code_toeslag: str) -> str:
    text = code_toeslag.lower()
    if "133" in text or "1.33" in text:
        return "133"
    if "135" in text or "1.35" in text:
        return "135"
    if "140" in text or "1.4" in text:
        return "140"
    if "180" in text or "1.8" in text:
        return "180"
    if "200" in text or "2.0" in text:
        return "200"
    if "300" in text or "3.0" in text:
        return "300"
    if "130" in text or "1.3" in text:
        return "130"
    return "100"


def _same_wagegroup(
    *,
    schaal_left: str | None,
    tarief_left: str | None,
    schaal_right: str | None,
    tarief_right: str | None,
) -> bool:
    return (schaal_left or "NA") == (schaal_right or "NA") and (tarief_left or "NA") == (
        tarief_right or "NA"
    )


def _pick_best_rate_card_candidate(
    *,
    invoice_rate: float,
    candidates: list["WagegroupRateCard"],
    tolerance_eur: float,
) -> "WagegroupRateCard | None":
    within_tolerance = [
        c for c in candidates if abs(float(invoice_rate) - float(c.rate_value)) <= tolerance_eur
    ]
    if not within_tolerance:
        return None
    within_tolerance.sort(key=lambda c: abs(float(invoice_rate) - float(c.rate_value)))
    return within_tolerance[0]


def _wagegroup_in_candidates(
    *,
    schaal: str | None,
    tarief: str | None,
    candidates: list["WagegroupRateCard"],
) -> bool:
    for c in candidates:
        if _same_wagegroup(
            schaal_left=schaal,
            tarief_left=tarief,
            schaal_right=c.schaal,
            tarief_right=c.tarief,
        ):
            return True
    return False


def _extract_schaal_tarief(text: str) -> tuple[str | None, str | None]:
    clean = _norm_text(text)
    if not clean:
        return None, None
    m = re.search(r"\b([A-Z]\d{1,2})\b", clean.upper())
    schaal = m.group(1) if m else None
    t = re.search(r"\bfase\s+([A-Z])\b", clean, flags=re.IGNORECASE)
    tarief = t.group(1).upper() if t else None
    return schaal, tarief


def _extract_phase(text: str) -> str | None:
    clean = _norm_text(text)
    if not clean:
        return None
    m = re.search(r"\bfase\s+([A-Z])\b", clean, flags=re.IGNORECASE)
    if m:
        return m.group(1).upper()
    if re.fullmatch(r"[A-Z]", clean.upper()):
        return clean.upper()
    return None


@dataclass
class ParsedRateWorkbook:
    person_rates: list[dict[str, Any]]
    rate_card: list[dict[str, Any]]


@dataclass(frozen=True)
class MappingRule:
    key: str
    header: str
    required: bool
    fallback_column: str | None = None


@dataclass
class PreviewContext:
    upload_id: str
    agency: str
    source_file_name: str
    workbook_path: str
    detected_sheet: str
    header_row: int
    header_cells: list[str]
    mapping: dict[str, dict[str, Any]]


_RATE_KEYS = ["100", "133", "135", "180", "200", "300"]
_ROW_ERROR_INLINE_LIMIT = 200
_SAMPLE_ROWS_LIMIT = 5
_SAMPLE_ROWS_MIN = 2

_AGENCY_CONFIG: dict[str, dict[str, Any]] = {
    "otto": {
        "sheets": ["Blad1"],
        "required_sheet": "Blad1",
        "mapping_rules": [
            MappingRule("person_number", "Personeelsnummer", True, None),
            MappingRule("first_name", "Voornaam", True, None),
            MappingRule("last_name", "Achternaam", True, None),
            MappingRule("tarief", "Fase tarief", False, "E"),
            MappingRule("schaal", "Schaal", False, "F"),
            MappingRule("rate_100", "100%", False, "Q"),
            MappingRule("rate_133", "133%", False, "R"),
            MappingRule("rate_135", "135%", False, "S"),
            MappingRule("rate_180", "180%", False, "T"),
            MappingRule("rate_200", "200%", False, "U"),
            MappingRule("rate_300", "300%", False, "V"),
        ],
    },
    "flexspecialisten": {
        "sheets": ["Blad1", "1 januari 2025"],
        "required_sheet": None,
        "mapping_rules": [
            MappingRule("person_number", "Loonnummer", True, None),
            MappingRule("first_name", "Voornaam", True, None),
            MappingRule("last_name", "Achternaam", True, None),
            MappingRule("loongroep", "Loongroep", False, "D"),
            MappingRule("fase_contracten", "Fase contracten", False, "E"),
            MappingRule("rate_100", "100%", False, "J"),
            MappingRule("rate_133", "133%", False, "L"),
            MappingRule("rate_135", "135%", False, "N"),
            MappingRule("rate_180", "180%", False, "V"),
            MappingRule("rate_200", "200%", False, "P"),
            MappingRule("rate_300", "300%", False, "R"),
        ],
    },
}


def _normalize_header(value: Any) -> str:
    text = _norm_text(value).replace("\ufeff", "")
    text = text.replace("％", "%")
    text = text.replace("procent", "%")
    text = re.sub(r"\s*%\s*", "%", text)
    text = re.sub(r"[\s,;:_-]+", " ", text.strip().lower())
    return " ".join(text.split())


def _column_letter_to_index(column: str) -> int | None:
    text = _norm_text(column).upper()
    if not text or not re.fullmatch(r"[A-Z]+", text):
        return None
    total = 0
    for ch in text:
        total = total * 26 + (ord(ch) - ord("A") + 1)
    return total - 1


def _column_index_to_letter(index: int) -> str | None:
    if index < 0:
        return None
    n = index + 1
    chars: list[str] = []
    while n > 0:
        n, rem = divmod(n - 1, 26)
        chars.append(chr(ord("A") + rem))
    return "".join(reversed(chars))


def _mapping_to_response_item(
    key: str,
    required: bool,
    mapping_entry: dict[str, Any],
) -> dict[str, Any]:
    return {
        "key": key,
        "header": mapping_entry.get("header"),
        "column": mapping_entry.get("column"),
        "sourceType": mapping_entry.get("sourceType", "missing"),
        "confidence": mapping_entry.get("confidence", "none"),
        "required": required,
        "notes": mapping_entry.get("notes"),
    }


def _select_sheet_for_agency(wb: Any, agency: str) -> Any:
    cfg = _AGENCY_CONFIG[agency]
    required_sheet = cfg.get("required_sheet")
    if required_sheet:
        if required_sheet not in wb.sheetnames:
            raise ValueError(
                f"{agency} tarievenbestand moet sheet '{required_sheet}' bevatten."
            )
        return wb[required_sheet]
    for candidate in cfg["sheets"]:
        if candidate in wb.sheetnames:
            return wb[candidate]
    raise ValueError(
        "Flex tarievenbestand moet sheet 'Blad1' of '1 januari 2025' bevatten."
    )


def _resolve_mapping(header_cells: list[str], agency: str) -> dict[str, dict[str, Any]]:
    by_exact: dict[str, int] = {}
    by_normalized: dict[str, int] = {}
    for idx, cell in enumerate(header_cells):
        if not cell:
            continue
        exact = _norm_text(cell)
        normalized = _normalize_header(cell)
        if exact and exact not in by_exact:
            by_exact[exact] = idx
        if normalized and normalized not in by_normalized:
            by_normalized[normalized] = idx

    mapping: dict[str, dict[str, Any]] = {}
    for rule in _AGENCY_CONFIG[agency]["mapping_rules"]:
        idx = by_exact.get(rule.header)
        if idx is not None:
            mapping[rule.key] = {
                "columnIndex": idx,
                "column": _column_index_to_letter(idx),
                "header": header_cells[idx] if idx < len(header_cells) else rule.header,
                "sourceType": "header",
                "confidence": "exact",
                "required": rule.required,
                "notes": None,
            }
            continue

        idx = by_normalized.get(_normalize_header(rule.header))
        if idx is not None:
            mapping[rule.key] = {
                "columnIndex": idx,
                "column": _column_index_to_letter(idx),
                "header": header_cells[idx] if idx < len(header_cells) else rule.header,
                "sourceType": "header",
                "confidence": "normalized",
                "required": rule.required,
                "notes": "Matched using normalized header value.",
            }
            continue

        if rule.fallback_column:
            fallback_idx = _column_letter_to_index(rule.fallback_column)
            if fallback_idx is not None:
                mapping[rule.key] = {
                    "columnIndex": fallback_idx,
                    "column": rule.fallback_column,
                    "header": (
                        header_cells[fallback_idx]
                        if fallback_idx < len(header_cells)
                        else None
                    ),
                    "sourceType": "fallback",
                    "confidence": "fallback",
                    "required": rule.required,
                    "notes": "Using agency fallback column.",
                }
                continue

        mapping[rule.key] = {
            "columnIndex": None,
            "column": None,
            "header": None,
            "sourceType": "missing",
            "confidence": "none",
            "required": rule.required,
            "notes": "No header or fallback column resolved.",
        }
    return mapping


def _compute_unresolved(mapping: dict[str, dict[str, Any]], agency: str) -> dict[str, list[str]]:
    unresolved_required: list[str] = []
    unresolved_optional: list[str] = []
    unresolved_rates: list[str] = []
    for rule in _AGENCY_CONFIG[agency]["mapping_rules"]:
        entry = mapping[rule.key]
        if entry.get("columnIndex") is not None:
            continue
        if rule.required:
            unresolved_required.append(rule.key)
        elif rule.key.startswith("rate_"):
            unresolved_rates.append(rule.key.replace("rate_", ""))
        else:
            unresolved_optional.append(rule.key)
    return {
        "unresolvedRequired": unresolved_required,
        "unresolvedOptional": unresolved_optional,
        "unresolvedRates": unresolved_rates,
    }


def _sample_rows_for_preview(ws: Any, mapping: dict[str, dict[str, Any]]) -> list[dict[str, Any]]:
    def _to_preview_value(value: Any) -> str | int | float | None:
        if value is None:
            return None
        if isinstance(value, bool):
            return str(value).lower()
        if isinstance(value, (int, float)):
            return value
        text = _norm_text(value)
        return text if text else None

    rows: list[dict[str, Any]] = []
    mapped_keys = [key for key, spec in mapping.items() if spec.get("columnIndex") is not None]
    if not mapped_keys:
        mapped_keys = list(mapping.keys())

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        row_sample: dict[str, Any] = {}
        has_value = False
        for key in mapped_keys:
            spec = mapping.get(key, {})
            col_idx = spec.get("columnIndex")
            raw_value = row[col_idx] if col_idx is not None and col_idx < len(row) else None
            value = _to_preview_value(raw_value)
            if value is not None:
                has_value = True
            row_sample[key] = value
        # Always include the first rows for deterministic UI preview,
        # then continue with non-empty mapped rows only.
        if has_value or len(rows) < _SAMPLE_ROWS_MIN:
            row_sample["rowNumber"] = row_idx
            rows.append(row_sample)
        if len(rows) >= _SAMPLE_ROWS_LIMIT:
            break
    return rows


def _persist_preview_context(preview_dir: str, context: PreviewContext) -> None:
    os.makedirs(preview_dir, exist_ok=True)
    context_path = os.path.join(preview_dir, f"{context.upload_id}.json")
    with open(context_path, "w", encoding="utf-8") as f:
        json.dump(asdict(context), f, ensure_ascii=True)


def _load_preview_context(preview_dir: str, upload_id: str) -> PreviewContext:
    context_path = os.path.join(preview_dir, f"{upload_id}.json")
    if not os.path.exists(context_path):
        raise ValueError("Onbekende uploadId; start met /upload_wagegroups_rate/preview.")
    with open(context_path, encoding="utf-8") as f:
        raw = json.load(f)
    return PreviewContext(**raw)


def create_wagegroup_rate_preview(
    *,
    content: bytes,
    filename: str,
    agency: str,
    preview_dir: str,
) -> dict[str, Any]:
    if load_workbook is None:
        raise ValueError("openpyxl is niet beschikbaar in deze omgeving.")
    wb = load_workbook(io.BytesIO(content), data_only=True)
    ws = _select_sheet_for_agency(wb, agency)

    header_cells = [_norm_text(c.value) for c in ws[1]]
    mapping = _resolve_mapping(header_cells, agency)
    unresolved = _compute_unresolved(mapping, agency)
    warnings: list[str] = []
    if unresolved["unresolvedRates"]:
        warnings.append(
            "Niet alle tariefkolommen konden worden gekoppeld: "
            + ", ".join(unresolved["unresolvedRates"])
        )

    upload_id = uuid.uuid4().hex
    workbook_path = os.path.join(preview_dir, f"{upload_id}.xlsx")
    os.makedirs(preview_dir, exist_ok=True)
    with open(workbook_path, "wb") as f:
        f.write(content)

    context = PreviewContext(
        upload_id=upload_id,
        agency=agency,
        source_file_name=filename,
        workbook_path=workbook_path,
        detected_sheet=str(ws.title),
        header_row=1,
        header_cells=header_cells,
        mapping=mapping,
    )
    _persist_preview_context(preview_dir, context)

    response_mapping = [
        _mapping_to_response_item(rule.key, rule.required, mapping[rule.key])
        for rule in _AGENCY_CONFIG[agency]["mapping_rules"]
    ]

    return {
        "uploadId": upload_id,
        "detectedSheet": ws.title if ws else None,
        "headerRow": 1,
        "mapping": response_mapping,
        "unresolvedRequired": unresolved["unresolvedRequired"],
        "unresolvedOptional": unresolved["unresolvedOptional"],
        "unresolvedRates": unresolved["unresolvedRates"],
        "warnings": warnings,
        "sampleRows": _sample_rows_for_preview(ws, mapping),
    }


def _apply_mapping_override(
    context: PreviewContext,
    mapping_override: dict[str, dict[str, str]] | None,
) -> tuple[dict[str, dict[str, Any]], list[str]]:
    merged = {k: dict(v) for k, v in context.mapping.items()}
    warnings: list[str] = []
    if not mapping_override:
        return merged, warnings

    header_lookup: dict[str, int] = {}
    normalized_lookup: dict[str, int] = {}
    for idx, cell in enumerate(context.header_cells):
        if not cell:
            continue
        header_lookup[cell] = idx
        normalized_lookup[_normalize_header(cell)] = idx

    for key, override in mapping_override.items():
        if key not in merged:
            warnings.append(f"Onbekende mappingOverride key genegeerd: {key}.")
            continue
        header_value = _norm_text(override.get("header")) if override else ""
        column_value = _norm_text(override.get("column")) if override else ""
        if not header_value and not column_value:
            # Ignore empty override entries and keep preview mapping.
            # Frontend may send sparse/default entries for untouched fields.
            continue

        idx = None
        if header_value:
            idx = header_lookup.get(header_value)
            if idx is None:
                idx = normalized_lookup.get(_normalize_header(header_value))
        if idx is None and column_value:
            idx = _column_letter_to_index(column_value)

        if idx is None:
            warnings.append(
                f"mappingOverride voor '{key}' genegeerd: geen geldige kolom/header gevonden."
            )
            continue

        merged[key] = {
            **merged[key],
            "columnIndex": idx,
            "column": _column_index_to_letter(idx),
            "header": context.header_cells[idx] if idx < len(context.header_cells) else None,
            "sourceType": "manual",
            "confidence": "manual",
            "notes": "Column set via mappingOverride.",
        }
    return merged, warnings


def _read_row_value(row: tuple[Any, ...], mapping: dict[str, dict[str, Any]], key: str) -> Any:
    idx = mapping.get(key, {}).get("columnIndex")
    if idx is None or idx >= len(row):
        return None
    return row[idx]


def commit_wagegroup_rate_preview(
    *,
    upload_id: str,
    agency: str,
    preview_dir: str,
    output_dir: str,
    mapping_override: dict[str, dict[str, str]] | None = None,
) -> tuple[ParsedRateWorkbook, dict[str, Any]]:
    if load_workbook is None:
        raise ValueError("openpyxl is niet beschikbaar in deze omgeving.")

    context = _load_preview_context(preview_dir, upload_id)
    if context.agency != agency:
        raise ValueError("uploadId hoort bij een andere agency.")
    if not os.path.exists(context.workbook_path):
        raise ValueError("Bronbestand voor uploadId is niet meer beschikbaar.")

    with open(context.workbook_path, "rb") as f:
        content = f.read()
    wb = load_workbook(io.BytesIO(content), data_only=True)
    ws = _select_sheet_for_agency(wb, agency)
    mapping, override_warnings = _apply_mapping_override(context, mapping_override)
    unresolved = _compute_unresolved(mapping, agency)
    if unresolved["unresolvedRequired"]:
        raise ValueError(
            "Ontbrekende verplichte kolommen na mappingOverride: "
            + ", ".join(unresolved["unresolvedRequired"])
            + ". Controleer mappingOverride.header/column."
        )

    errors: list[dict[str, Any]] = []
    warnings: list[str] = list(override_warnings)
    person_rates: list[dict[str, Any]] = []
    rate_card_candidates: dict[tuple[str, str, str], list[float]] = defaultdict(list)
    processed_rows = 0
    ingested_rows = 0
    skipped_rows = 0

    for row_number, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        processed_rows += 1
        person_number = _to_person_number(_read_row_value(row, mapping, "person_number"))
        first_name = _norm_text(_read_row_value(row, mapping, "first_name"))
        last_name = _norm_text(_read_row_value(row, mapping, "last_name"))
        full_name = _norm_text(f"{first_name} {last_name}") or (person_number or "")

        row_errors: list[dict[str, Any]] = []
        if not person_number:
            row_errors.append(
                {"row": row_number, "code": "missing_person_number", "message": "Persoonsnummer ontbreekt.", "field": "person_number"}
            )
        if not first_name or not last_name:
            row_errors.append(
                {"row": row_number, "code": "missing_name", "message": "Voornaam en/of achternaam ontbreekt.", "field": "name"}
            )
        if row_errors:
            skipped_rows += 1
            errors.extend(row_errors)
            continue

        if agency == "otto":
            tarief = _extract_phase(_read_row_value(row, mapping, "tarief"))
            schaal = _norm_text(_read_row_value(row, mapping, "schaal")) or None
            provider = "otto"
        else:
            loongroep = _norm_text(_read_row_value(row, mapping, "loongroep"))
            schaal, parsed_tarief = _extract_schaal_tarief(loongroep)
            fase = _extract_phase(_read_row_value(row, mapping, "fase_contracten"))
            tarief = fase or parsed_tarief or "NA"
            schaal = schaal or "NA"
            provider = "flexspecialisten"

        valid_rate_count = 0
        for rate_key in _RATE_KEYS:
            key = f"rate_{rate_key}"
            if mapping.get(key, {}).get("columnIndex") is None:
                continue
            raw_rate = _read_row_value(row, mapping, key)
            rate_value = _to_float(raw_rate)
            if raw_rate not in (None, "") and rate_value is None:
                errors.append(
                    {
                        "row": row_number,
                        "code": "invalid_rate_value",
                        "message": f"Ongeldige numerieke waarde voor {rate_key}%.",
                        "field": key,
                    }
                )
                continue
            if rate_value is None:
                continue
            valid_rate_count += 1
            person_rates.append(
                {
                    "provider": provider,
                    "person_number": person_number,
                    "name": full_name or person_number,
                    "normalized_name": normalize_person_name(full_name or person_number),
                    "schaal": schaal,
                    "tarief": tarief,
                    "rate_key": rate_key,
                    "rate_value": rate_value,
                    "source_file": context.source_file_name,
                    "source_week": None,
                }
            )
            rate_card_candidates[(str(schaal or "NA"), str(tarief or "NA"), rate_key)].append(
                rate_value
            )

        if valid_rate_count == 0:
            skipped_rows += 1
            errors.append(
                {
                    "row": row_number,
                    "code": "no_rates_found",
                    "message": "Geen valide tariefwaarden gevonden in de rij.",
                    "field": "rates",
                }
            )
            continue
        ingested_rows += 1

    rate_card: list[dict[str, Any]] = []
    for (schaal, tarief, rate_key), values in rate_card_candidates.items():
        selected = Counter(round(v, 4) for v in values).most_common(1)[0][0]
        rate_card.append(
            {
                "provider": agency,
                "schaal": schaal,
                "tarief": tarief,
                "rate_key": rate_key,
                "rate_value": float(selected),
                "source_file": context.source_file_name,
                "source_week": None,
            }
        )

    report_url: str | None = None
    inline_errors: list[dict[str, Any]] | None = errors
    if len(errors) > _ROW_ERROR_INLINE_LIMIT:
        os.makedirs(output_dir, exist_ok=True)
        report_path = os.path.join(output_dir, f"{upload_id}_wagegroup_rate_errors.json")
        with open(report_path, "w", encoding="utf-8") as f:
            json.dump(errors, f, ensure_ascii=True, indent=2)
        report_url = report_path
        inline_errors = errors[:_ROW_ERROR_INLINE_LIMIT]
        warnings.append("Foutenrapport is ingekort inline; volledig rapport staat op errorReportUrl.")

    stats = {
        "processedRows": processed_rows,
        "ingestedRows": ingested_rows,
        "skippedRows": skipped_rows,
        "errorReportUrl": report_url,
        "errorReportInline": inline_errors,
        "warnings": warnings,
    }
    return ParsedRateWorkbook(person_rates=person_rates, rate_card=rate_card), stats


def parse_otto_wagegroup_rate_workbook(
    content: bytes, filename: str
) -> ParsedRateWorkbook:
    if load_workbook is None:
        raise ValueError("openpyxl is niet beschikbaar in deze omgeving.")
    wb = load_workbook(io.BytesIO(content), data_only=True)
    if "Blad1" not in wb.sheetnames:
        raise ValueError("OTTO tarievenbestand moet sheet 'Blad1' bevatten.")
    ws = wb["Blad1"]

    header = [(_norm_text(c.value)) for c in ws[1]]
    by_name: dict[str, int] = {}
    for i, h in enumerate(header):
        if not h:
            continue
        key = h.lower()
        if key not in by_name:
            by_name[key] = i
    person_idx = by_name.get("personeelsnummer")
    first_name_idx = by_name.get("voornaam")
    last_name_idx = by_name.get("achternaam")
    tarief_idx = by_name.get("fase tarief")
    schaal_idx = by_name.get("schaal")
    if tarief_idx is None:
        tarief_idx = 4  # E
    if schaal_idx is None:
        schaal_idx = 5  # F
    if person_idx is None:
        raise ValueError("Kolom 'Personeelsnummer' is verplicht.")
    if first_name_idx is None or last_name_idx is None:
        raise ValueError("Kolommen 'Voornaam' en 'Achternaam' zijn verplicht.")

    rate_columns = [
        ("100%", "100"),
        ("133%", "133"),
        ("135%", "135"),
        ("180%", "180"),
        ("200%", "200"),
        ("300%", "300"),
    ]
    # Fallback positions for OTTO sheet layout when headers are missing:
    # Q,R,S,T,U,V -> 100,133,135,180,200,300
    fallback_rate_indexes = {
        "100": 16,  # Q
        "133": 17,  # R
        "135": 18,  # S
        "180": 19,  # T
        "200": 20,  # U
        "300": 21,  # V
    }
    rate_indexes: list[tuple[int, str]] = []
    for col_name, rate_key in rate_columns:
        idx = by_name.get(col_name)
        if idx is None:
            idx = fallback_rate_indexes[rate_key]
        rate_indexes.append((idx, rate_key))
    if not rate_indexes:
        raise ValueError(
            "Geen OTTO tariefkolommen gevonden (verwacht 100%/133%/135%/180%/200%/300%)."
        )

    person_rates: list[dict[str, Any]] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        try:
            person_number = _to_person_number(
                row[person_idx] if person_idx < len(row) else None
            )
            if not person_number:
                continue
            first_name = _norm_text(
                row[first_name_idx] if first_name_idx < len(row) else None
            )
            last_name = _norm_text(
                row[last_name_idx] if last_name_idx < len(row) else None
            )
            name = _norm_text(f"{first_name} {last_name}")
            tarief = _extract_phase(row[tarief_idx] if tarief_idx < len(row) else None)
            schaal = (
                _norm_text(row[schaal_idx] if schaal_idx < len(row) else None) or None
            )
            for idx, rate_key in rate_indexes:
                rate_value = _to_float(row[idx] if idx < len(row) else None)
                if rate_value is None:
                    continue
                person_rates.append(
                    {
                        "provider": "otto",
                        "person_number": person_number,
                        "name": name or person_number,
                        "normalized_name": normalize_person_name(name or person_number),
                        "schaal": schaal,
                        "tarief": tarief,
                        "rate_key": rate_key,
                        "rate_value": rate_value,
                        "source_file": filename,
                        "source_week": None,
                    }
                )
        except Exception:
            continue

    # derive card as median-ish mode by (schaal, tarief, key)
    grouped: dict[tuple[str, str, str], list[float]] = defaultdict(list)
    for row in person_rates:
        schaal = row.get("schaal") or ""
        tarief = row.get("tarief") or ""
        grouped[(str(schaal), str(tarief), str(row["rate_key"]))].append(
            float(row["rate_value"])
        )
    rate_card: list[dict[str, Any]] = []
    for (schaal, tarief, rate_key), values in grouped.items():
        selected = Counter(round(v, 4) for v in values).most_common(1)[0][0]
        rate_card.append(
            {
                "provider": "otto",
                "schaal": schaal or "NA",
                "tarief": tarief or "NA",
                "rate_key": rate_key,
                "rate_value": float(selected),
                "source_file": filename,
                "source_week": None,
            }
        )
    return ParsedRateWorkbook(person_rates=person_rates, rate_card=rate_card)


def parse_flex_wagegroup_rate_workbook(
    content: bytes, filename: str
) -> ParsedRateWorkbook:
    if load_workbook is None:
        raise ValueError("openpyxl is niet beschikbaar in deze omgeving.")
    wb = load_workbook(io.BytesIO(content), data_only=True)
    if "Blad1" in wb.sheetnames:
        ws = wb["Blad1"]
    elif "1 januari 2025" in wb.sheetnames:
        ws = wb["1 januari 2025"]
    else:
        raise ValueError(
            "Flex tarievenbestand moet sheet 'Blad1' of '1 januari 2025' bevatten."
        )

    header = [(_norm_text(c.value)) for c in ws[1]]
    by_name: dict[str, int] = {h.lower(): i for i, h in enumerate(header) if h}
    person_idx = by_name.get("loonnummer")
    first_name_idx = by_name.get("voornaam")
    last_name_idx = by_name.get("achternaam")
    loongroep_idx = by_name.get("loongroep")
    fase_contracten_idx = by_name.get("fase contracten")
    if loongroep_idx is None:
        loongroep_idx = 3  # D
    if fase_contracten_idx is None:
        fase_contracten_idx = 4  # E
    if person_idx is None:
        raise ValueError("Kolom 'Loonnummer' is verplicht.")
    if first_name_idx is None or last_name_idx is None:
        raise ValueError("Kolommen 'Voornaam' en 'Achternaam' zijn verplicht.")

    rate_columns = [
        ("100%", "100"),
        ("133%", "133"),
        ("135%", "135"),
        ("180%", "180"),
        ("200%", "200"),
        ("300%", "300"),
    ]
    # Fallback positions for Flex layout when % headers are missing.
    # J,L,N,P,R -> 100,133,135,200,300
    fallback_rate_indexes = {
        "100": 9,  # J
        "133": 11,  # L
        "135": 13,  # N
        "200": 15,  # P
        "300": 17,  # R
    }
    rate_indexes: list[tuple[int, str]] = []
    for col_name, rate_key in rate_columns:
        idx = by_name.get(col_name)
        if idx is None:
            idx = fallback_rate_indexes.get(rate_key)
        if idx is not None:
            rate_indexes.append((idx, rate_key))
    if not rate_indexes:
        raise ValueError(
            "Geen Flex tariefkolommen gevonden (verwacht 100%/133%/135%/180%/200%/300%)."
        )

    person_rates: list[dict[str, Any]] = []
    rate_card_candidates: dict[tuple[str, str, str], list[float]] = defaultdict(list)
    for row in ws.iter_rows(min_row=2, values_only=True):
        try:
            person_number = _to_person_number(
                row[person_idx] if person_idx < len(row) else None
            )
            if not person_number:
                continue
            first_name = _norm_text(
                row[first_name_idx] if first_name_idx < len(row) else None
            )
            last_name = _norm_text(
                row[last_name_idx] if last_name_idx < len(row) else None
            )
            name = _norm_text(f"{first_name} {last_name}")

            loongroep = _norm_text(
                row[loongroep_idx] if loongroep_idx < len(row) else None
            )
            schaal, parsed_tarief = _extract_schaal_tarief(loongroep)
            fase_contracten = _norm_text(
                row[fase_contracten_idx] if fase_contracten_idx < len(row) else None
            )
            fase = _extract_phase(fase_contracten)
            tarief = fase or parsed_tarief or "NA"
            schaal_value = schaal or "NA"

            for idx, rate_key in rate_indexes:
                rate_value = _to_float(row[idx] if idx < len(row) else None)
                if rate_value is None:
                    continue
                person_rates.append(
                    {
                        "provider": "flexspecialisten",
                        "person_number": person_number,
                        "name": name or person_number,
                        "normalized_name": normalize_person_name(name or person_number),
                        "schaal": schaal_value,
                        "tarief": tarief,
                        "rate_key": rate_key,
                        "rate_value": rate_value,
                        "source_file": filename,
                        "source_week": None,
                    }
                )
                rate_card_candidates[(schaal_value, tarief, rate_key)].append(
                    rate_value
                )
        except Exception:
            continue

    rate_card: list[dict[str, Any]] = []
    for (schaal, tarief, rate_key), values in rate_card_candidates.items():
        selected = Counter(round(v, 4) for v in values).most_common(1)[0][0]
        rate_card.append(
            {
                "provider": "flexspecialisten",
                "schaal": schaal,
                "tarief": tarief,
                "rate_key": rate_key,
                "rate_value": float(selected),
                "source_file": filename,
                "source_week": None,
            }
        )
    return ParsedRateWorkbook(person_rates=person_rates, rate_card=rate_card)


async def persist_parsed_wagegroup_rates(
    db: AsyncSession,
    *,
    provider: str,
    parsed: ParsedRateWorkbook,
) -> dict[str, int]:
    await db.execute(
        delete(PersonWagegroupRate).where(PersonWagegroupRate.provider == provider)
    )
    await db.execute(
        delete(WagegroupRateCard).where(WagegroupRateCard.provider == provider)
    )
    db.add_all(PersonWagegroupRate(**row) for row in parsed.person_rates)
    db.add_all(WagegroupRateCard(**row) for row in parsed.rate_card)
    await db.commit()
    return {
        "personRates": len(parsed.person_rates),
        "rateCardRows": len(parsed.rate_card),
    }


def write_rates_csvs(
    *, provider: str, parsed: ParsedRateWorkbook, output_dir: str
) -> dict[str, str]:
    os.makedirs(output_dir, exist_ok=True)
    person_path = os.path.join(output_dir, f"{provider}_wagegroup_person_rates.csv")
    card_path = os.path.join(output_dir, f"{provider}_wagegroup_rate_card.csv")
    with open(person_path, "w", newline="", encoding="utf-8-sig") as f:
        fields = [
            "provider",
            "person_number",
            "name",
            "normalized_name",
            "schaal",
            "tarief",
            "rate_key",
            "rate_value",
            "source_file",
            "source_week",
        ]
        w = csv.DictWriter(f, fieldnames=fields)
        w.writeheader()
        w.writerows(parsed.person_rates)
    with open(card_path, "w", newline="", encoding="utf-8-sig") as f:
        fields = [
            "provider",
            "schaal",
            "tarief",
            "rate_key",
            "rate_value",
            "source_file",
            "source_week",
        ]
        w = csv.DictWriter(f, fieldnames=fields)
        w.writeheader()
        w.writerows(parsed.rate_card)
    return {"personRatesCsv": person_path, "rateCardCsv": card_path}


def _invoice_line_rate(row: InvoiceLine) -> float | None:
    if row.totaal_uren and row.totaal_uren > 0:
        return row.subtotaal / row.totaal_uren
    if row.uurloon and row.uurloon > 0:
        return row.uurloon
    return None


def _write_rate_mismatches_csv(path: str, rows: list[dict[str, Any]]) -> None:
    with open(path, "w", newline="", encoding="utf-8-sig") as f:
        fields = [
            "sapId",
            "name",
            "invoiceCodeToeslag",
            "rateKey",
            "invoiceRate",
            "expectedRate",
            "difference",
            "status",
            "matchMethod",
        ]
        w = csv.DictWriter(f, fieldnames=fields)
        w.writeheader()
        w.writerows(rows)


def _build_histogram_rows(diffs: list[float]) -> list[dict[str, Any]]:
    bins = [
        (0, 0.1, "0.00-0.10"),
        (0.1, 0.25, "0.10-0.25"),
        (0.25, 0.5, "0.25-0.50"),
        (0.5, 1.0, "0.50-1.00"),
        (1.0, 2.0, "1.00-2.00"),
        (2.0, 5.0, "2.00-5.00"),
        (5.0, 999999.0, "5.00+"),
    ]
    rows = []
    for lo, hi, label in bins:
        count = sum(1 for d in diffs if lo <= d < hi)
        rows.append({"bucket": label, "count": count})
    return rows


def _write_histogram_csv(path: str, diffs: list[float]) -> None:
    rows = _build_histogram_rows(diffs)
    with open(path, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.DictWriter(f, fieldnames=["bucket", "count"])
        w.writeheader()
        w.writerows(rows)


def _write_wagegroup_rate_differences_csv(path: str, rows: list[dict[str, Any]]) -> None:
    with open(path, "w", newline="", encoding="utf-8-sig") as f:
        fields = [
            "provider",
            "personId",
            "name",
            "invoiceCodeToeslag",
            "rateKey",
            "invoiceRate",
            "deducedInvoiceWagegroup",
            "referenceWagegroup",
            "difference",
            "status",
            "matchMethod",
        ]
        w = csv.DictWriter(f, fieldnames=fields)
        w.writeheader()
        w.writerows(rows)


async def analyze_wagegroup_differences_by_rate(
    *,
    week: int,
    provider: str,
    db: AsyncSession,
    tolerance_eur: float,
    output_dir: str,
) -> dict[str, Any]:
    person_rates_result = await db.execute(
        select(PersonWagegroupRate).where(PersonWagegroupRate.provider == provider)
    )
    person_rates = list(person_rates_result.scalars().all())
    card_result = await db.execute(
        select(WagegroupRateCard).where(WagegroupRateCard.provider == provider)
    )
    card_rows = list(card_result.scalars().all())
    invoice_result = await db.execute(
        select(InvoiceLine).where(
            InvoiceLine.week_number == week,
            InvoiceLine.agency == provider,
        )
    )
    invoice_rows = list(invoice_result.scalars().all())

    person_lookup: dict[tuple[str, str], PersonWagegroupRate] = {}
    person_name_lookup: dict[tuple[str, str], PersonWagegroupRate] = {}
    for row in person_rates:
        person_lookup[(row.person_number, row.rate_key)] = row
        person_name_lookup[(row.normalized_name, row.rate_key)] = row

    card_by_rate_key: dict[str, list[WagegroupRateCard]] = defaultdict(list)
    for row in card_rows:
        card_by_rate_key[row.rate_key].append(row)

    mismatches: list[dict[str, Any]] = []
    matched_rows = 0
    missing_reference = 0
    missing_rate_card_match = 0

    for row in invoice_rows:
        rate_key = _rate_key_from_code_toeslag(row.code_toeslag or "")
        invoice_rate = _invoice_line_rate(row)
        if invoice_rate is None:
            continue

        sap_id = _norm_text(row.sap_id)
        normalized_name = normalize_person_name(row.naam or "")

        person_rate = person_lookup.get((sap_id, rate_key))
        match_method = "person_number"
        if not person_rate:
            person_rate = person_name_lookup.get((normalized_name, rate_key))
            match_method = "name_fallback"
        if not person_rate:
            missing_reference += 1
            continue

        reference_wagegroup = f"{person_rate.schaal or 'NA'} / Fase {person_rate.tarief or 'NA'}"
        card_candidates = card_by_rate_key.get(rate_key, [])
        deduced_candidates = [
            c
            for c in card_candidates
            if abs(float(invoice_rate) - float(c.rate_value)) <= tolerance_eur
        ]
        deduced = _pick_best_rate_card_candidate(
            invoice_rate=float(invoice_rate),
            candidates=deduced_candidates,
            tolerance_eur=tolerance_eur,
        )
        if not deduced:
            missing_rate_card_match += 1
            continue
        matched_rows += 1
        if _wagegroup_in_candidates(
            schaal=person_rate.schaal,
            tarief=person_rate.tarief,
            candidates=deduced_candidates,
        ):
            continue

        deduced_wagegroup = f"{deduced.schaal or 'NA'} / Fase {deduced.tarief or 'NA'}"
        reference_rate_diff = abs(float(invoice_rate) - float(person_rate.rate_value))

        mismatches.append(
            {
                "provider": provider,
                "personId": sap_id or person_rate.person_number,
                "name": row.naam or person_rate.name,
                "invoiceCodeToeslag": row.code_toeslag,
                "rateKey": rate_key,
                "invoiceRate": round(float(invoice_rate), 4),
                "deducedInvoiceWagegroup": deduced_wagegroup,
                "referenceWagegroup": reference_wagegroup,
                "difference": round(reference_rate_diff, 4),
                "status": "wagegroup_diff_by_rate",
                "matchMethod": match_method,
            }
        )

    os.makedirs(output_dir, exist_ok=True)
    path = os.path.join(output_dir, f"{week} validation_wagegroups_by_rate_{provider}.csv")
    _write_wagegroup_rate_differences_csv(path, mismatches)
    return {
        "status": "ok",
        "week": week,
        "provider": provider,
        "toleranceEur": tolerance_eur,
        "matchedRows": matched_rows,
        "missingReference": missing_reference,
        "missingRateCardMatch": missing_rate_card_match,
        "wagegroupDifferences": len(mismatches),
        "mismatches": mismatches,
        "outputFile": path,
    }


async def analyze_otto_rate_mismatches(
    *,
    week: int,
    db: AsyncSession,
    tolerance_eur: float,
    output_dir: str,
) -> dict[str, Any]:
    person_rates_result = await db.execute(
        select(PersonWagegroupRate).where(PersonWagegroupRate.provider == "otto")
    )
    person_rates = list(person_rates_result.scalars().all())
    card_result = await db.execute(
        select(WagegroupRateCard).where(WagegroupRateCard.provider == "otto")
    )
    card_rows = list(card_result.scalars().all())
    known_result = await db.execute(
        select(PersonWagegroup).where(PersonWagegroup.provider == "otto")
    )
    known_rows = list(known_result.scalars().all())
    known_by_person = {r.person_number: r for r in known_rows}
    known_by_name = {normalize_person_name(r.name): r for r in known_rows}

    person_rate_lookup: dict[tuple[str, str], PersonWagegroupRate] = {}
    person_rate_name_lookup: dict[tuple[str, str], PersonWagegroupRate] = {}
    for row in person_rates:
        person_rate_lookup[(row.person_number, row.rate_key)] = row
        person_rate_name_lookup[(row.normalized_name, row.rate_key)] = row

    card_lookup: dict[tuple[str, str, str], WagegroupRateCard] = {}
    for row in card_rows:
        card_lookup[(row.schaal, row.tarief, row.rate_key)] = row

    invoice_result = await db.execute(
        select(InvoiceLine).where(
            InvoiceLine.week_number == week,
            InvoiceLine.agency == "otto",
        )
    )
    invoice_rows = list(invoice_result.scalars().all())

    mismatches: list[dict[str, Any]] = []
    differences: list[float] = []
    matched = 0
    missing_expected_rate = 0
    for row in invoice_rows:
        rate_key = _rate_key_from_code_toeslag(row.code_toeslag or "")
        invoice_rate = _invoice_line_rate(row)
        if invoice_rate is None:
            continue
        sap_id = _norm_text(row.sap_id)
        normalized_name = normalize_person_name(row.naam or "")
        known = known_by_person.get(sap_id) or known_by_name.get(normalized_name)
        person_rate = person_rate_lookup.get((sap_id, rate_key))
        match_method = "person_number"
        if not person_rate:
            person_rate = person_rate_name_lookup.get((normalized_name, rate_key))
            match_method = "name_fallback"
        expected_rate = person_rate.rate_value if person_rate else None
        if expected_rate is None and known:
            # fallback via card using known wagegroup as schaal+tarief hint
            schaal, tarief = _extract_schaal_tarief(known.wagegroup)
            if schaal and tarief:
                card = card_lookup.get((schaal, tarief, rate_key))
                if card:
                    expected_rate = card.rate_value
                    match_method = "rate_card"
        if expected_rate is None:
            missing_expected_rate += 1
            continue
        matched += 1
        diff = abs(invoice_rate - expected_rate)
        differences.append(diff)
        if diff <= tolerance_eur:
            continue
        mismatches.append(
            {
                "sapId": sap_id,
                "name": row.naam,
                "invoiceCodeToeslag": row.code_toeslag,
                "rateKey": rate_key,
                "invoiceRate": round(invoice_rate, 4),
                "expectedRate": round(expected_rate, 4),
                "difference": round(diff, 4),
                "status": "rate_mismatch",
                "matchMethod": match_method,
            }
        )

    os.makedirs(output_dir, exist_ok=True)
    mismatch_path = os.path.join(output_dir, f"{week} validation_wagerates_otto.csv")
    hist_path = os.path.join(output_dir, f"{week} rate_diff_histogram_otto.csv")
    _write_rate_mismatches_csv(mismatch_path, mismatches)
    _write_histogram_csv(hist_path, differences)

    return {
        "status": "ok",
        "week": week,
        "provider": "otto",
        "toleranceEur": tolerance_eur,
        "matchedRateRows": matched,
        "missingExpectedRate": missing_expected_rate,
        "rateMismatches": len(mismatches),
        "mismatches": mismatches,
        "outputFile": mismatch_path,
        "histogramFile": hist_path,
    }
