import csv
import os
import re
from datetime import datetime, date

import openpyxl


def get_week_prefix(wb):
    """
    Scan sheets for a 'Datum' column and return a 'YYYYww ' prefix derived from
    the first data value found. Returns '' if no Datum column is found.
    Dutch week numbering == ISO 8601, so isocalendar() is correct.
    """
    for sn in wb.sheetnames:
        ws = wb[sn]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        header = [str(c).strip() if c is not None else "" for c in rows[0]]
        if "Datum" not in header:
            continue
        col_idx = header.index("Datum")
        for row in rows[1:]:
            val = row[col_idx] if col_idx < len(row) else None
            if val is None or val == "":
                continue
            if isinstance(val, (datetime, date)):
                d = val
            else:
                try:
                    d = datetime.strptime(str(val).strip()[:10], "%Y-%m-%d")
                except ValueError:
                    continue
            iso = d.isocalendar()  # (year, week, weekday)
            return f"{iso[0]}{iso[1]:02d} "
    return ""


def _to_safe_base(fname: str) -> str:
    base = re.sub(r"\.xlsx$", "", fname, flags=re.IGNORECASE)
    base = base.replace(",", " ")
    return " ".join(base.split()).strip()


def _convert_workbook(fpath: str, output_dir: str) -> list[str]:
    fname = os.path.basename(fpath)
    safe_base = _to_safe_base(fname)

    wb = openpyxl.load_workbook(fpath, data_only=True)
    # Only prepend the prefix if the filename doesn't already start with one (YYYYww )
    prefix = "" if re.match(r"^\d{6} ", safe_base) else get_week_prefix(wb)
    output_base = f"{prefix}{safe_base}".strip()

    created_files: list[str] = []
    # Generic sheet names that don't add useful information to the filename
    generic_sheet_names = {"sheet", "sheet1", "blad1", "blad"}
    is_factuur_file = "specificatie" in output_base.lower()
    for sn in wb.sheetnames:
        ws = wb[sn]
        safe_sn = sn.replace("/", "-").replace("\\", "-").strip()
        is_generic = safe_sn.lower() in generic_sheet_names
        should_append_sheet_name = (len(wb.sheetnames) > 1 and not is_generic) or (
            is_factuur_file and safe_sn.lower() == "export factuur"
        )
        if should_append_sheet_name:
            out_name = f"{output_base} - {safe_sn}.csv"
        else:
            out_name = f"{output_base}.csv"
        out_path = os.path.join(output_dir, out_name)
        with open(out_path, "w", newline="", encoding="utf-8-sig") as csvf:
            writer = csv.writer(csvf)
            for row in ws.iter_rows(values_only=True):
                writer.writerow([v if v is not None else "" for v in row])
        created_files.append(out_path)
        print(f"Created: {out_name}")

    return created_files


def convert_input(kloklijst_filename, factuur_filename):
    """Convert only the two provided XLSX files into CSV in formatted_input/."""

    input_dir = "input"
    output_dir = "formatted_input"
    os.makedirs(output_dir, exist_ok=True)

    candidates = [kloklijst_filename, factuur_filename]
    created_files: list[str] = []

    for candidate in candidates:
        fpath = candidate
        if not os.path.isabs(fpath):
            fpath = os.path.join(input_dir, candidate)
        fname = os.path.basename(fpath)

        if fname.startswith("~$"):
            continue
        if not fname.lower().endswith(".xlsx"):
            raise ValueError(f"Alleen .xlsx-bestanden worden ondersteund: {fname}")
        if not os.path.exists(fpath):
            raise FileNotFoundError(f"Invoerbestand niet gevonden: {fpath}")

        created_files.extend(_convert_workbook(fpath, output_dir))

    return created_files


if __name__ == "__main__":
    raise SystemExit(
        "Gebruik convert_input(kloklijst_filename, factuur_filename) vanuit de applicatiecode."
    )
